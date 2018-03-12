# -*- coding: utf-8 -*-
# plugin (local modules) import

from general_helpers import datestr_to_python
from general_helpers import highlight_submenu
from general_helpers import get_badge
from general_helpers import get_label
from general_helpers import get_last_day_month
from general_helpers import get_months_list
from general_helpers import class_get_teachers
from general_helpers import get_paused_subscriptions
from general_helpers import max_string_length
from general_helpers import Memo_links
from general_helpers import set_form_id_and_get_submit_button
#from general_helpers import workshops_get_full_workshop_product_id

from os_storage import uploads_available_space

from openstudio import *

# from openstudio import \
#     Classcard, \
#     ClasscardsHelper, \
#     AttendanceHelper, \
#     WorkshopsHelper, \
#     Customer, \
# from os_customers import CustomersHelper
# from os_customer_subscriptions import CustomerSubscriptionsHelper
# from os_customer_subscriptions import CustomerSubscription
# from os_school_subscriptions import SchoolSubscription
# from os_school_classcards import SchoolClasscard
# from os_invoices import Invoice
# from os_invoices import InvoicesHelper
# from os_orders import Order

# python general modules import
import cStringIO
import os
import openpyxl
import calendar
import codecs
#TODO: change all titles to "Customer"
#TODO: change all subtitles to customer name


# helper functions

def _edit_check_picture(form):
    if form.vars['Picture'] == '':
        row = db.auth_user[form.vars['id']]
        row.picture = None
        row.thumbsmall = None
        row.thumblarge = None
        row.update_record()

def classes_check_reservation(row):
    '''
        Check if a customer is already reserved for a class.
        If not, return an add button to use in the grid.
    '''
    customers_id = request.vars['cuID']
    clsID = row.id
    check = db.classes_reservation(customers_id=customers_id, classes_id=clsID)
    if check is None:
        return os_gui.get_button('add', URL("classes", "reserve_class",
                                     vars={'cuID':customers_id,
                                           'clsID':clsID}))
                                           #'back'="customers_classes")
    else:
        return ""

#TODO: Rename & reuse for subscriptions_alt_prices repeat
@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('update', 'alternativepayments'))
def alternativepayment_repeat():
    apID = request.vars['apID']
    row = db.alternativepayments[apID]

    year = row.PaymentYear
    month = row.PaymentMonth

    if month == 12:
        month = 1
        year += 1
    else:
        month += 1
    db.alternativepayments.insert(auth_customer_id=row.auth_customer_id,
                                  PaymentYear=year,
                                  PaymentMonth=month,
                                  payment_categories_id=row.payment_categories_id,
        Amount=row.Amount, Description=row.Description)

    session.customers_payments_tab = '#ap'

    redirect(URL("payments",vars={'cuID':row.auth_customer_id}))


def index_get_export(val=None):
    '''
        Returns dict with export button and bs3 modal containing the links
        to different export options.
    '''
    mailinglist = A((os_gui.get_fa_icon('fa-envelope-o'),
                     T("Mailing list")),
                    _href=URL('export_excel', vars=dict(export='mailing_list')),
                    _class='textalign_left')
    active_customers = A((os_gui.get_fa_icon('fa-check'),
                         T("Active customers")),
                         _href=URL('export_excel',
                                   vars=dict(export='customers_list')),
                         _class='textalign_left')

    links = [ mailinglist, active_customers ]

    export = os_gui.get_dropdown_menu(
            links = links,
            btn_text = '',
            btn_icon = 'download',
            btn_size = 'btn-sm',
            menu_class='pull-right' )

    return export


def _edit_clear_old_thumbs(form):
    """
        This function cleans up generated thumbnails of the old picture when a new picture is uploaded.
        It does nothing when no new picture is uploaded.
    """
    customers_id = session.customers_edit_clear_thumbs_customers_id
    row = db.customers[customers_id]

    if not form.vars['Picture'] is None:
        if not row.ThumbSmall is None:
            os.remove(request.folder + 'uploads/' + row.ThumbSmall)
        if not row.ThumbLarge is None:
            os.remove(request.folder + 'uploads/' + row.ThumbLarge)


def edit_remodel_form(form,
                      picture,
                      change_picture,
                      label_id,
                      customers_id,
                      bo_button='',
                      te_button='',
                      mail_button='',
                      merged=False,
                      row=None):
    """
        This function takes the default form and makes a nice looking desktop browser friendly form out of it.
        row = db.auth_user(customers_id)
    """

    # check for contact permissions
    contact_permission = auth.has_membership(group_id='Admins') or \
                         auth.has_permission('update', 'customers_contact')

    if not contact_permission:
        # hide contact info
        form.custom.label.phone = ''
        form.custom.widget.phone = ''
        form.custom.label.mobile = ''
        form.custom.widget.mobile = ''
        form.custom.label.email = ''
        form.custom.widget.email = ''
        form.custom.label.company = ''
        form.custom.widget.company = ''
        form.custom.label.emergency = ''
        form.custom.widget.emergency = ''

    # check for address permissions
    address_permission = auth.has_membership(group_id='Admins') or \
                         auth.has_permission('update', 'customers_address')

    if not address_permission:
        # hide contact info
        form.custom.label.address = ''
        form.custom.widget.address = ''
        form.custom.label.postcode = ''
        form.custom.widget.postcode = ''
        form.custom.label.city = ''
        form.custom.widget.city = ''
        form.custom.label.country = ''
        form.custom.widget.country = ''

    div_picture = DIV(LABEL(form.custom.label.newsletter),
                      form.custom.widget.newsletter,
                      BR(),
                      picture,
                      change_picture, ' ',
                      SPAN(label_id, ' ', customers_id,
                           _id='customers_id'),
                      _class='col-md-6')

    # basic info
    div_basic_info = DIV( TABLE(
            TR( TD(LABEL(form.custom.label.first_name)),
                TD(form.custom.widget.first_name),
                ),
            TR( TD(LABEL(form.custom.label.last_name)),
                TD(form.custom.widget.last_name)
                ),
            TR( TD(LABEL(form.custom.label.date_of_birth)),
                TD(form.custom.widget.date_of_birth),
                ),
            TR( TD(LABEL(form.custom.label.gender)),
                TD(form.custom.widget.gender),
                ),
            TR( TD(LABEL(form.custom.label.email)),
                TD(DIV(form.custom.widget.email,
                       DIV(mail_button,
                            _class='input-group-btn'),
                       _class='input-group')),
                ),
            TR( TD(LABEL(form.custom.label.phone)),
                TD(form.custom.widget.phone),
                ),
            TR( TD(LABEL(form.custom.label.mobile)),
                TD(form.custom.widget.mobile),
                ),
            TR( TD(LABEL(form.custom.label.company)),
                TD(form.custom.widget.company),
                ),
            TR( TD(LABEL(form.custom.label.emergency)),
                TD(form.custom.widget.emergency),
                ),
            _class='full-width'),
            _class='col-md-6 customers_edit_basic_info os-no_margin_bottom')



    # check if we have to separate customers by location
    if session.show_location:
        location_label = LABEL(form.custom.label.school_locations_id)
        location_widget = form.custom.widget.school_locations_id
    else:
        location_label = ''
        location_widget = ''

    bo_label = ''
    bo_note = ''
    bo_all_notes = ''
    te_label = ''
    te_note = ''
    te_all_notes = ''
    if not customers_id is None and not customers_id == '':
        if auth.has_membership(group_id='Admins') or \
           auth.has_permission('read', 'customers_notes_backoffice'):
            bo_label = LABEL(T("Notes"), BR(), T("Back office"))
            bo_note = DIV(LOAD('customers', 'notes.load', ajax=True,
                                        target='os-bonote_latest',
                                        vars={'cuID':customers_id,
                                              'note_type':'backoffice',
                                              'latest':True,
                                              'latest_length':140}),
                                  _class='os-customers_note_latest')

        if auth.has_membership(group_id='Admins') or \
           auth.has_permission('read', 'customers_notes_teachers'):
            te_label = LABEL(T("Notes"), BR(), T("Teachers"))
            te_note = DIV(LOAD('customers', 'notes.load', ajax=True,
                                     target='os-tenote_latest',
                                   vars={'cuID':customers_id,
                                         'note_type':'teachers',
                                         'latest':True,
                                         'latest_length':140}),
                                _class='os-customers_note_latest')

    # address info
    div_address = DIV(
        H3(T("Address and studio info")),
        DIV(
            TABLE(
                TR( TD(LABEL(form.custom.label.address)),
                    TD(form.custom.widget.address),
                    ),
                TR( TD(LABEL(form.custom.label.city)),
                    TD(form.custom.widget.city),
                    ),
                TR( TD(LABEL(form.custom.label.school_levels_id)),
                    TD(form.custom.widget.school_levels_id),
                    ),
                TR( TD(LABEL(form.custom.label.keynr)),
                    TD(form.custom.widget.keynr),
                    ),
                TR( TD(location_label),
                    TD(location_widget)),
                _class='full-width'),
        _class='col-md-6'),
        DIV(TABLE(
            TR( TD(LABEL(form.custom.label.postcode)),
                TD(form.custom.widget.postcode)),
            TR( TD(LABEL(form.custom.label.country)),
                TD(form.custom.widget.country)),
            TR( TD(LABEL(form.custom.label.school_discovery_id)),
                TD(form.custom.widget.school_discovery_id)),
            TR( TD(LABEL(form.custom.label.school_languages_id)),
                TD(form.custom.widget.school_languages_id)),
            _class='full-width'),
            _class='col-md-6'),
        _class='col-md-12 customers_edit_address_info')

    notes = DIV(
        DIV(
            TABLE( TR( TD(bo_label),
                       TD(bo_note, bo_button)),
                  _class='full-width'),
            _class='col-md-6'),
        DIV(
            TABLE( TR( TD(te_label),
                        TD(te_note, te_button)),
                  _class='full-width'),
            _class='col-md-6'),
        _class='col-md-12 customers_edit_notes')

    return DIV(XML('<form id="MainForm" action="#" enctype="multipart/form-data" method="post">'),
               DIV(div_picture,
                   div_basic_info,
                   _class='col-md-12'),
               div_address,
               notes,
               form.custom.end,
               _class='customers_edit_container row')


# def index_get_subscription_query(query, mstype, msmonth, msyear):
#     msstartdate = datetime.date(msyear, msmonth, 1)
#     next_month = msstartdate.replace(day=28) + datetime.timedelta(days=4)  # this will never fail
#     msenddate = next_month - datetime.timedelta(days=next_month.day)
#     if mstype != '':
#         query &= (db.customers_subscriptions.Startdate <= msenddate) & \
#                  ((db.customers_subscriptions.Enddate >= msstartdate) | \
#                   (db.customers_subscriptions.Enddate == None)) & \
#             (db.customers_subscriptions.auth_customer_id == db.auth_user.id) & \
#             (db.customers_subscriptions.school_subscriptions_id==mstype)
#             # inner join for search :
#             #(db.customers_subscriptions.customers_id == db.customers.id) & \
#     return query


def subscriptions_get_link_latest_pauses(row):
    '''
        Returns latest pauses for a subscription
    '''
    csID = row.id
    cuID = row.auth_customer_id
    query = (db.customers_subscriptions_paused.customers_subscriptions_id == row.id)
    rows = db(query).select(db.customers_subscriptions_paused.ALL,
                            orderby=~db.customers_subscriptions_paused.Startdate,
                            limitby=(0,3))

    pause_list = DIV()
    for row in rows:
        item = SPAN(row.Startdate, ' - ', row.Enddate, ' ',
                    _title=row.Description,
                    _class='grey small_font')
        pause_list.append(item)
        pause_list.append(BR())

    pause_list.append(A(SPAN(T("Edit pauses"), _class='small_font'),
                        _href=URL('subscription_pauses', vars={'cuID':cuID,
                                                               'csID':csID}),
                        _title=T("View all pauses and add new")))

    return pause_list


def subscriptions_get_link_credits(row):
    '''
        Returns total number of credits for a subscription
    '''
    cs = CustomerSubscription(row.id)

    credits = cs.get_credits_balance()

    return A(credits,
             _href=URL('subscription_credits', vars={'cuID':row.auth_customer_id,
                                                     'csID':row.id}))


### helpers end ###

@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'auth_user'))
def index():
    # make sure we're redirected back to the list from the edit page
    session.customers_back = None
    # Redirect back to edit page after adding
    session.customers_add_back = None

    response.search_available = True
    try:
        response.q = session.customers_load_list_search_name.replace('%', '')

    except AttributeError:
        response.q = ''


    # archive filter
    show = 'current'

    if 'show_archive' in request.vars:
        show = request.vars['show_archive']
        session.customers_show = show

    if not session.customers_show:
        session.customers_show = 'current'

    if session.customers_show == 'archive':
        archive_class = 'active'
        current_class = ''
    else:
        current_class = 'active'
        archive_class = ''

    if 'nr_items' in request.vars:
        session.customers_index_items_per_page = int(request.vars['nr_items'])

    archive_buttons = os_gui.get_archived_radio_buttons(session.customers_show)

    response.title = T("Customers")

    export = ''
    if auth.has_membership(group_id='Admins') or auth.has_permission('update', 'auth_user'):
        export = index_get_export()


    add = ''
    if ( auth.has_membership(group_id='Admins') or
         auth.has_permission('create', 'auth_user') ):
        ch = CustomersHelper()
        result = ch.get_add_modal()
        add = SPAN(result['button'], result['modal'], _class='pull-right')

    if session.customers_show == 'current':
        archived = False
    else:
        archived = True

    show_location = False
    if session.show_location:
        show_location = 'True'

    show_email = False
    if ( auth.has_membership(group_id='Admins') or
         auth.has_permission('update', 'customer-contact') ):
        show_email = True


    search_results = DIV(LOAD('customers', 'load_list.load',
                              target='customers_load_list',
                              content=os_gui.get_ajax_loader(message=T("Searching...")),
                              vars={'list_type':'customers_index',
                                    'items_per_page':session.customers_index_items_per_page,
                                    'initial_list':True,
                                    'archived':archived,
                                    'show_location':show_location,
                                    'show_email': show_email},
                              ajax=True),
                         _id="customers_load_list",
                         _class="load_list_customers clear")

    # archive_buttons = os_gui.get_archived_radio_buttons(
    #     session.customers_show)

    content = DIV(
        UL(LI(A(T('Current'),
                _href=URL(vars={'show_archive':'current'})),
              _class=current_class),
           LI(A(T('Archive'),
                _href=URL(vars={'show_archive':'archive'})),
              _class=archive_class),
           # LI(I(_class='fa fa-users'),
           #    _class='pull-left header'),
           _class='nav nav-tabs pull-right'),
        DIV(DIV(search_results,
                _class='tab-pane active'),
            _class='tab-content'),
        _class='nav-tabs-custom')


    tools = index_get_tools()


    return dict(add=add,
                export=export,
                content=content,
                nr_items=index_get_select_nr_items(),
                header_tools=tools)


def index_get_link_archive(row):
    '''
        Called from the index function. Changes title of archive button
        depending on whether a customer is archived or not
    '''
    row = db.auth_user(row.id)

    try:
        if row.archived:
            tt = T("Move to current")
        else:
            tt = T("Archive")

        return os_gui.get_button('archive',
                                 URL('archive',
                                     vars={'uID':row.id},
                                     extension=''),
                                 tooltip=tt)
    except AttributeError: # might get thrown if a customer is deleted, but still in cache. Then the row can't be fetched
        return


def index_get_select_nr_items(var=None):
    '''
        Returns a form to select number of items to show on a page
    '''
    view_set = [10, 15, 25]
    form = SQLFORM.factory(
        Field('nr_items',
              requires=IS_IN_SET(view_set, zero=None),
              default=session.customers_index_items_per_page or 10,
              label=T('Customers per page')),
        formstyle='divs',
    )

    select = form.element('#no_table_nr_items')
    select['_onchange'] = 'this.form.submit();'
    select['_class'] += ' inline-block'

    form = DIV(DIV(form.custom.begin,
                   form.custom.label.nr_items, ' ',
                   form.custom.widget.nr_items,
                   form.custom.end,
                   _class='well pull-right'),
               _id='customers_index_settings',
               _class='collapse')

    return form


def index_get_tools(var=None):
    '''
        Returns tools menu for customers list
    '''
    tools = []

    # teacher holidays
    permission = auth.has_membership(group_id='Admins') or \
                 auth.has_permission('read', 'customers_subscriptions_credits')

    if permission:
        subscription_credits = A(os_gui.get_fa_icon('fa-check-square-o'),
                                 T("Subscription credits"),
                                 _href=URL('customers', 'subscription_credits_month'),
                                 _title=T('List subscription credits for a selected month'))
        tools.append(subscription_credits)

    # get menu
    tools = os_gui.get_dropdown_menu(tools,
                                     '',
                                     btn_size='btn-sm',
                                     btn_icon='wrench',
                                     menu_class='pull-right'
                                     )

    return tools


@auth.requires(auth.has_membership(group_id='Admins') or
               auth.has_permission('update', 'auth_user'))
def archive():
    '''
        Archive an account
    '''
    uID = request.vars['uID']
    if not uID:
        session.flash = T('Unable to (un)archive customer')
    else:
        row = db.auth_user(uID)

        if row.archived:
            session.flash = T('Moved to current')
        else:
            # set enddate for recurring reservations
            query = (db.classes_reservation.auth_customer_id == uID) & \
                    (db.classes_reservation.ResType == 'recurring') & \
                    ((db.classes_reservation.Enddate == None) |
                     (db.classes_reservation.Enddate > TODAY_LOCAL))
            db(query).update(Enddate = TODAY_LOCAL)
            # remove all waitinglist entries
            query = (db.classes_waitinglist.auth_customer_id == uID)
            db(query).delete()

            session.flash = T('Archived')


        row.archived = not row.archived
        row.update_record()

    redirect(URL('index'))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('delete', 'auth_user'))
def delete():
    '''
        Delete a customer
    '''
    cuID = request.vars['cuID']

    query = (db.auth_user.id == cuID)
    db(query).update(trashed=True)

    session.flash = T('Deleted')

    redirect(URL('index'))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'auth_user'))
def add():
    '''
        Page to add a new customer, only show the required field and after
        adding redirect to the edit page
    '''
    # call js for styling the form
    response.js = 'set_form_classes();'

    # enable only required fields
    for field in db.auth_user:
        field.readable=False
        field.writable=False

    db.auth_user.first_name.readable=True
    db.auth_user.first_name.writable=True
    db.auth_user.last_name.readable=True
    db.auth_user.last_name.writable=True
    db.auth_user.email.readable=True
    db.auth_user.email.writable=True

    if session.show_location:
        db.auth_user.school_locations_id.readable = True
        db.auth_user.school_locations_id.writable = True

        db.auth_user.school_locations_id.requires = \
            IS_IN_DB(db,
                     'school_locations.id',
                     '%(Name)s',
                     zero=T("Please select..."))

    db.auth_user.password.default = generate_password(30)
    db.auth_user.customer.default = True

    if request.vars['teacher'] == 'True':
        db.auth_user.teacher.default = True
        db.auth_user.login_start.default = 'backend'
        crud.settings.create_onaccept = [cache_clear_school_teachers]

    if request.vars['employee'] == 'True':
        db.auth_user.employee.default = True
        db.auth_user.login_start.default  = 'backend'

    if 'clsID' in request.vars:
        # if we're comming from classes/attendance_list
        clsID = request.vars['clsID']
        date_formatted  = request.vars['date']
        next_url = '/customers/add_redirect_on_create?cuID=[id]'
        next_url += '&clsID=' + clsID
        next_url += '&date='  + date_formatted
    else:
        next_url = '/customers/add_redirect_on_create?cuID=[id]'

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Saved")
    crud.settings.create_next = next_url
    form = crud.create(db.auth_user)

    form_id = "customer_add"
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    submit = form.element('input[type=submit]')

    # Make table inputs full width
    table = form.element('table')
    table['_class'] = 'full-width'

    return dict(content=form)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('create', 'auth_user'))
def add_redirect_on_create():
    '''
        Redirect to edit, from the client side, to leave the add modal
    '''
    cuID = request.vars['cuID']

    if session.customers_add_back == 'classes_attendance':
        session.flash = T("Added customer")
        redirect(URL('classes', 'attendance', vars=request.vars),
                 client_side=True)
    elif session.customers_add_back == 'school_employees':
        session.flash = T("Added employee")
        redirect(URL('school_properties', 'employees'),
                 client_side=True)
    elif session.customers_add_back == 'school_teachers':
        session.flash = T("Added teacher")
        redirect(URL('school_properties', 'teachers'),
                 client_side=True)
    else:
        # to edit page
        redirect(URL('edit', args=[cuID], extension=''), client_side=True)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('update', 'auth_user'))
def edit():
    today = TODAY_LOCAL

    if not session.show_location:
        db.auth_user.school_locations_id.readable = False
        db.auth_user.school_locations_id.writable = False
    else:
        db.auth_user.school_locations_id.requires = \
            IS_IN_DB(db,
                     'school_locations.id',
                     '%(Name)s',
                     zero=T("Please select..."))

    db.auth_user.picture.readable=False
    db.auth_user.picture.writable=False

    # we're not entering password, so don't include it in the form
    db.auth_user.password.readable=False
    db.auth_user.password.writable=False


    picture_class = 'customer_image_edit'
    picture = DIV(IMG(_src=URL('static', 'images/person.png'), _alt="Person picture"),
                      _class=picture_class)

    # update customer
    customers_id = request.args[0]
    db.auth_user.id.label = T('Customer ID')

    row = db.auth_user[customers_id]
    response.title = row.display_name
    response.subtitle = T("Profile")
    if row.merged:
        menu = ''
    else:
        menu = customers_get_menu(customers_id, 'general')

    left_sidebar_enabled = True
    label_id = T('ID') + ": "

    if not row.picture:
        change_picture_title = T("Add picture")
        picture = DIV(IMG(_src=URL('static', 'images/person.png'),
                          _alt=row.display_name),
                          _class=picture_class)
    else:
        picture = DIV(IMG(_src=URL('default', 'download',
                                   args=row.thumblarge),
                          _alt=row.display_name),
                          _class=picture_class)
        change_picture_title = T("Change picture")
    change_picture = A(change_picture_title,
                       _href=URL('edit_picture', args=[customers_id]))

    #crud.settings.update_onaccept.auth_user.append(_check_active)
    crud.messages.submit_button = T('Save')
    crud.messages.record_updated = T('Saved')

    # Clear teachers cache if we're updating a teacher
    if row.teacher:
        crud.settings.update_onaccept = [cache_clear_school_teachers]

    form = crud.update(db.auth_user, customers_id)
    # Tie the elements together using the form html5 attribute.
    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = "MainForm"

    submit = os_gui.get_submit_button('MainForm')

    back = edit_get_back()

    # set mail button
    email = row.email
    mail_button = A(I(_class='fa fa-envelope-o'),
                      _class="btn btn-default",
                      _href='mailto:' + email)

    # add notes
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'customers_notes_backoffice'):
        bo_notes = LOAD('customers', 'notes.load', ajax=True,
                        vars={'cuID':customers_id,
                              'note_type':'backoffice'})

        bo_result = os_gui.get_modal(button_text=T("All notes"),
                                    modal_title=T('Back office notes'),
                                    modal_content=bo_notes,
                                    modal_class='customers_bo_notes',
                                    button_id='all_bo_notes',
                                    button_class='btn-xs pull-right')
        bo_modal = bo_result['modal']
        bo_button = bo_result['button']

    else:
        bo_button =''
        bo_modal = ''

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'customers_notes_teachers'):
        te_notes = LOAD('customers', 'notes.load', ajax=True,
                        vars={'cuID':customers_id,
                              'note_type':'teachers'})

        te_result = os_gui.get_modal(button_text=T("All notes"),
                                     modal_title=T('Teacher notes'),
                                     modal_content=te_notes,
                                     modal_class='customers_te_notes',
                                     button_class='btn-xs pull-right')
        te_modal = te_result['modal']
        te_button = te_result['button']

    else:
        te_button =''
        te_modal = ''

    # get styles form
    form = edit_remodel_form(form,
                             picture,
                             change_picture,
                             label_id,
                             customers_id,
                             bo_button=bo_button,
                             te_button=te_button,
                             mail_button=mail_button,
                             merged=row.merged,
                             row=row)

    alert = ''
    if row.merged:
        merged_into = db.auth_user(row.merged_into)
        merged_link = A(SPAN(merged_into.display_name,
                             T('ID'), ': ',
                             row.merged_into),
                        _title=T("link to account merged into"),
                        _href=URL('edit', args=[row.merged_into]))
        alert = os_gui.get_alert(
            'success',
            SPAN(B(T('Note')), ' ',
                 T("This account has been merged into"), ' ',
                 merged_link, ' ', T('on'), ' ',
                 row.merged_on.strftime(DATETIME_FORMAT), '.'),
            dismissable=False)
        # Don't show a submit button for merged customers
        submit = ''

    content = DIV(bo_modal,
                  te_modal,
                  alert,
                  form,
                  _class = 'tab-pane active')


    return dict(content=content,
                back=back,
                menu=menu,
                save=submit)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'auth_user'))
def edit_picture():
    """
        This function is called to show a page to change a customers' picture
        After successfully changing the picture, the user is redirected to the edit page.
    """
    response.title = T("Picture")
    response.view = 'general/only_content.html'
    customers_id = request.args[0]
    row = db.auth_user[customers_id]
    response.subtitle = row.display_name

    session.customers_edit_clear_thumbs_customers_id = customers_id

    for field in db.auth_user:
        field.writable = False
        field.readable = False

    db.auth_user.picture.readable = True
    db.auth_user.picture.writable = True
    db.auth_user.picture.label = T("Picture")

    submit = ''
    space = uploads_available_space(request.folder)
    if space['available'] > 1 or not row.Picture is None:
        crud.settings.label_separator = ''
        crud.settings.update_deletable = False
        crud.settings.update_onvalidation.customers.append(_edit_clear_old_thumbs)
        crud.settings.update_onaccept.customers.extend([_edit_check_picture, cache_clear_school_teachers])

        crud.messages.submit_button = T('Save')
        crud.messages.record_updated = T('Saved')

        form = crud.update(db.auth_user, customers_id, next='edit/[id]')

        result = set_form_id_and_get_submit_button(form, 'MainForm')
        form = result['form']
        submit = result['submit']

        msg_list = UL(T("Maximum filesize: 4MB"))

        content = DIV(BR(), msg_list, form)
    else:
        content = space['full_message']

    back = os_gui.get_button('back', URL('edit', args=[customers_id]))

    return dict(content=content, save=submit, back=back)


def edit_get_back(_class=''):
    '''
        This function looks at the session variable
            session.customers_back
        to determine where the back button for customers edit pages should link to.
    '''
    if session.customers_back == 'keys':
        # check if we came from the edit button on the keys list page in school properties
        url = URL('school_properties', 'list_keys')
    elif session.customers_back == 'teachers':
        # check if we came from the school / teachers page
        url = URL('school_properties', 'teachers')
    elif session.customers_back == 'school_employees':
        # check if we came fromthe school/employees page
        url = URL('school_properties', 'employees')
    elif session.customers_back == 'pinboard':
        # check if the birthday notification or a memo on the pinboard
        # referred to this page
        url = URL('pinboard', 'index')
    elif session.customers_back == 'subscriptions_new' or \
       session.customers_back == 'subscriptions_stopped' or \
       session.customers_back == 'subscriptions_paused' or \
       session.customers_back == 'subscriptions_overview_customers' or \
       session.customers_back == 'subscriptions_alt_prices':
        # check if the we came from the default/subscriptions page
        url = URL('reports', session.customers_back)
    elif session.customers_back == 'trialclasses':
        # check if the we came from the default/trialclasses page
        url = URL('reports', 'trialclasses')
    elif session.customers_back == 'trialcards':
        # check if the we came from the default/trialcards page
        url = URL('reports', 'trialcards')
    elif session.customers_back == 'dropinclasses':
        # check if the we came from the default/dropinclasses page
        url = URL('reports', 'dropinclasses')
    elif session.customers_back == 'classcards':
        # check if the we came from the default/classcards page
        url = URL('reports', 'classcards')
    elif session.customers_back == 'direct_debit_extra':
        # check if the we came from the default/alternative_payments page
        url = URL('reports', 'direct_debit_extra')
    elif session.customers_back == 'reports_attendance_subscription_exceeded':
        # check if we came from the reports/attendance_subcription_exceeded page
        url = URL('reports', 'attendance_subcription_exceeded')
    elif session.customers_back == 'reports_retention_rate':
        # check if we came from the reports/attendance_subcription_exceeded page
        url = URL('reports', 'retention_rate')
    # elif session.customers_back == 'classes_manage':
    #     # check if the classes/manage page referred to this page
    #     url = URL('classes', 'manage')
    # elif session.customers_back == 'workshops_attendance_add':
    #     # check if the workshops manage page referred to this page
    #     url = URL('workshops', 'attendance_add')
    # elif session.customers_back == 'workshops_attendance':
    #     # check if the workshops attendance page referred to this page
    #     url = URL('workshops', 'attendance')
    # elif session.customers_back == 'workshops_activity_attendance':
    #     # check if the workshops manage page referred to this page
    #     url = URL('workshops', 'activity_attendance')
    # elif session.customers_back == 'workshops_activity_attendance_add':
    #     # check if the workshops manage page referred to this page
    #     url = URL('workshops', 'activity_attendance_add')
    elif session.customers_back == 'finance_orders':
        # check if the finance orders page referred to this page
        url = URL('finance', 'orders')
    elif session.customers_back == 'tasks_index':
        # check if the tasks index page referred to this page
        url = URL('tasks', 'index')
    elif session.customers_back == 'finance_batch_content':
        # check if we're coming from a batch content page
        url = URL('finance', 'batch_content')
    else:
        url = URL('index')

    return os_gui.get_button('back', url, _class=_class)


def customers_get_menu(customers_id, page=None):
    pages = []

    pages.append(['general',
                   T('Profile'),
                  URL("customers","edit", args=[customers_id])])
    auth_user = db.auth_user(customers_id)
    if auth_user.teacher:
        pages.append(['edit_teacher',
                      T('Teacher profile'),
                      URL('edit_teacher', vars={'cuID':customers_id})])

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'customers_subscriptions'):
        pages.append(['subscriptions',
                      T("Subscriptions"),
                      URL("customers","subscriptions",
                          vars={'cuID':customers_id})])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'customers_classcards'):
        pages.append(['classcards',
                      T("Class cards"),
                      URL("customers","classcards", vars={'cuID':customers_id})])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'classes_reservation'):
        pages.append(['classes',
                      T("Classes"),
                      URL("customers", "classes_attendance",
                          vars={'cuID':customers_id})])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'workshops'):
        pages.append(['events',
                      T("Events"),
                      URL("customers","events", vars={'cuID':customers_id})])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'customers_documents'):
        pages.append(['documents',
                      T("Documents"),
                      URL("customers","documents", vars={'cuID':customers_id})])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'tasks'):
        pages.append(['tasks',
                      T("Tasks"),
                      URL("customers","tasks", vars={'cuID':customers_id})])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'orders'):
        pages.append(['orders',
                      T('Orders'),
                      URL("customers","orders", vars={'cuID':customers_id})])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'invoices'):
        pages.append(['invoices',
                      T("Invoices"),
                      URL("customers","invoices", vars={'cuID':customers_id})])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'customers_payments'):
        pages.append(['payments',
                      T("Payment info"),
                      URL("customers", "payments", vars={'cuID':customers_id})])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'auth_user_account'):
        pages.append(['account',
                      T('Account'),
                      URL('customers', 'account', vars={'cuID':customers_id})])

    return os_gui.get_submenu(pages, page, _id='os-customers_edit_menu', horizontal=True, htype='tabs')


def classcards_count_classes(row):
    ccd = Classcard(row.customers_classcards.id)
    link_text = ccd.get_classes_remaining_formatted()


    # card = db.school_classcards(row.customers_classcards.school_classcards_id)
    # total = card.Classes
    # ccdID = row.customers_classcards.id
    # if card.Unlimited:
    #     remaining = T('Unlimited')
    # else:
    #     query = (db.classes_attendance.customers_classcards_id == ccdID)
    #     used = db(query).count()
    #     remaining = total - used
    #
    # if remaining != 1:
    #     link_text = T("Classes")
    # else:
    #     link_text = T("Class")

    link = A(link_text,
             _href=URL('classcard_classes', vars=dict(ccdID=row.customers_classcards.id)))

    return link



@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'auth_user'))
def export_excel():
    """
    valid export_types include: payment_summary, mailinglist, attendance_list, customers_list, subscription_list
    """
    export_type = request.vars['export']

    # create filestream
    stream = cStringIO.StringIO()

    if export_type.lower() == "customers_list":
        # create dictionary to lookup latest subscription values
        if export_type.lower() == 'customers_list':
            result = db.executesql("""SELECT cu.id,
                                             ssu.name,
                                             cs.startdate,
                                             cs.enddate,
                                             pm.Name
                                       FROM auth_user cu
                LEFT JOIN customers_subscriptions cs
                ON cs.auth_customer_id = cu.id
                LEFT JOIN
                (SELECT auth_customer_id, school_subscriptions_id, max(startdate) as startdate, enddate
                FROM customers_subscriptions GROUP BY auth_customer_id) chk
                ON cu.id = chk.auth_customer_id
                LEFT JOIN
                (SELECT id, name FROM school_subscriptions) ssu
                ON ssu.id = cs.school_subscriptions_id
                LEFT JOIN payment_methods pm ON cs.payment_methods_id = pm.id
                WHERE cs.startdate = chk.startdate OR cs.startdate IS NULL """)

            fname = T("Customers list") + '.xlsx'
            title = 'Customers list'
        cu_mem_dict = dict()
        for record in result:
            subscription = record[1] or ""
            start = record[2] or ""
            end = record[3] or ""
            payment_method = record[4] or ""
            cu_mem_dict[record[0]] = [subscription, start, end, payment_method]

        bd_dict = dict()
        rows = db().select(db.customers_payment_info.ALL)
        for row in rows:
            bd_dict[row.auth_customer_id] = [
                                          row.AccountNumber,
                                          row.AccountHolder,
                                          row.BankName,
                                          row.BankLocation ]

        # Create the workbook
        wb = openpyxl.workbook.Workbook(write_only=True)
        ws = wb.create_sheet(title=title)
        headers = [ "id",
                    "First name",
                    "Last name",
                    "Date of birth",
                    "Gender",
                    "Address",
                    "Postal code",
                    "City",
                    "Country",
                    "Email",
                    "Newsletter",
                    "Telephone",
                    "Mobile",
                    "Key",
                    "Location",
                    "Subscription",
                    "Startdate",
                    "Enddate",
                    "Payment",
                    "AccountNR",
                    "AccountHolder",
                    "BankName",
                    "BankLocation"]
        ws.append(headers)

        query = (db.auth_user.archived == False)
        rows = db(query).select(db.auth_user.ALL,
                                db.school_locations.Name,
                left=[db.school_locations.on(db.auth_user.school_locations_id==\
                                             db.school_locations.id)])
        for row in rows:
            customers_id = row.auth_user.id
            if export_type.lower() == 'subscription_list' and \
                not cu_mem_dict.has_key(customers_id):
                # subscription list, if no subscription --> check the next customer.
                continue
            else:
                data = [ row.auth_user.id,
                         row.auth_user.first_name,
                         row.auth_user.last_name,
                         row.auth_user.date_of_birth,
                         row.auth_user.gender,
                         row.auth_user.address,
                         row.auth_user.postcode,
                         row.auth_user.city,
                         row.auth_user.country,
                         row.auth_user.email,
                         row.auth_user.newsletter,
                         row.auth_user.phone,
                         row.auth_user.mobile,
                         row.auth_user.keynr,
                         row.school_locations.Name,
                         cu_mem_dict[customers_id][0],
                         cu_mem_dict[customers_id][1],
                         cu_mem_dict[customers_id][2],
                         cu_mem_dict[customers_id][3]]
                if not bd_dict.get(customers_id, None) is None:
                    data.append(bd_dict[customers_id][0])
                    data.append(bd_dict[customers_id][1])
                    data.append(bd_dict[customers_id][2])
                    data.append(bd_dict[customers_id][3])

                ws.append(data)

        wb.save(stream)

        response.headers['Content-Type']='application/vnd.ms-excel'
        response.headers['Content-disposition']='attachment; filename=' + fname

        return stream.getvalue()


    elif export_type.lower() == "mailing_list":
        wb = openpyxl.workbook.Workbook(write_only=True)
        # write the sheet for all mail addresses
        ws = wb.create_sheet(title="All customers")
        today = datetime.date.today()
        query = ((db.auth_user.archived == False) &
                 (db.auth_user.id > 1))
        rows = db(query).select(db.auth_user.first_name,
                                db.auth_user.last_name,
                                db.auth_user.email)
        for row in rows:
            ws.append([row.first_name,
                       row.last_name,
                       row.email])
        # All newsletter
        ws = wb.create_sheet(title="Newsletter")
        today = datetime.date.today()
        query = ((db.auth_user.archived == False) &
                 (db.auth_user.newsletter == True) &
                 (db.auth_user.id > 1))
        rows = db(query).select(db.auth_user.first_name,
                                db.auth_user.last_name,
                                db.auth_user.email)
        for row in rows:
            ws.append([row.first_name,
                       row.last_name,
                       row.email])

        fname = T("Mailinglist") + '.xlsx'
        wb.save(stream)

        response.headers['Content-Type']='application/vnd.ms-excel'
        response.headers['Content-disposition']='attachment; filename=' + fname

        return stream.getvalue()


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_classcards'))
def classcards():
    '''
        List class cards for a customer
    '''
    response.view = 'customers/edit_general.html'
    customers_id = request.vars['cuID']

    customer = Customer(customers_id)
    response.title = customer.get_name()
    response.subtitle = T("Class cards")

    session.invoices_payment_add_back = 'customers_classcards'

    query = (db.customers_classcards.auth_customer_id == customers_id)
    db.customers_classcards.id.label = T("Pass number")
    db.customers_classcards.Enddate.readable = True

    maxtextlengths = {'customers_classcards.school_classcards_id' : 32,
                      'customers_classcards.Note' : 20}

    left = [
        db.invoices_customers_classcards.on(
            db.invoices_customers_classcards.customers_classcards_id ==
            db.customers_classcards.id),
        db.invoices.on(db.invoices_customers_classcards.invoices_id ==
                       db.invoices.id) ]

    links = [ dict(header=T('Classes'), body=classcards_count_classes),
              dict(header=T('Invoices'), body=classcards_get_link_invoice),
              lambda row: os_gui.get_button('edit',
                    URL('classcard_edit',
                        vars={'cuID':customers_id,
                              'ccdID':row.customers_classcards.id}))]
    fields = [ db.customers_classcards.id,
               db.customers_classcards.Startdate,
               db.customers_classcards.Enddate,
               db.customers_classcards.school_classcards_id,
               db.customers_classcards.Note,
               db.invoices.id,
               db.invoices.Status,
               db.invoices.payment_methods_id ]
    headers = { 'customers_classcards.id':T("Card") }

    delete_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('delete', 'customers_classcards')

    db.invoices.id.readable = False
    db.invoices.Status.readable = False

    grid = SQLFORM.grid(query,
                        fields=fields,
                        headers=headers,
                        field_id=db.customers_classcards.id,
                        left=left,
                        links=links,
                        maxtextlengths=maxtextlengths,
                        create=False,
                        details=False,
                        csv=False,
                        searchable=False,
                        editable=False,
                        deletable = delete_permission,
                        ondelete=classcards_ondelete,
                        orderby=~db.customers_classcards.id,
                        ui=grid_ui)
    grid.element('.web2py_counter', replace=None) # remove the counter
    grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    add_url = URL('classcard_add', vars={'cuID':customers_id})
    add = os_gui.get_button('add', add_url, T("Add a new class card"), btn_size='btn-sm', _class='pull-right')

    back = edit_get_back()

    menu = customers_get_menu(customers_id, request.function)

    return dict(content=grid, menu=menu, back=back, add=add)


def classcards_ondelete(table, record_id):
    '''
        Function to clear cache for customer after deleting a card 
    '''
    # Find customer id
    ccd = db.customers_classcards(record_id)
    cuID = ccd.auth_customer_id

    # Cancel invoice(s) for this classcard
    query = (db.invoices_customers_classcards.customers_classcards_id == record_id)
    rows = db(query).select(db.invoices_customers_classcards.ALL)
    for row in rows:
        invoice_query = (db.invoices.id == row.invoices_id)
        db(invoice_query).update(Status='cancelled')

    # Clear cache
    cache_clear_customers_classcards(cuID)


def classcards_get_link_invoice(row):
    '''
        Returns invoice for classcard in list
    '''
    if row.invoices.id:
        ih = InvoicesHelper()

        query = (db.invoices.id == row.invoices.id)
        rows = db(query).select(db.invoices.ALL)
        repr_row = rows.render(0)

        invoice_link = ih.represent_invoice_for_list(
            row.invoices.id,
            repr_row.InvoiceID,
            repr_row.Status,
            row.invoices.Status,
            row.invoices.payment_methods_id
        )

    else:
        invoice_link = ''

    return invoice_link


def classcards_get_return_url(cuID, clsID=None, date_formatted=None):
    '''
        returns the return url for class cards
    '''
    if clsID:
        url = URL('classes', 'attendance_list_classcards',
                     vars={'cuID' : cuID,
                           'clsID': clsID,
                           'date' : date_formatted})
    else:
        url = URL('classcards', vars={'cuID':cuID})

    return url


def classcards_clear_cache(form):
    '''
        Clear the subscriptions cache for customer 
    '''
    ccdID = form.vars.id
    ccd = db.customers_classcards(ccdID)
    cuID = ccd.auth_customer_id

    cache_clear_customers_classcards(cuID)


@auth.requires_login()
def classcard_add():
    '''
        Determine whether to use the classic or new style classcards add page
        > 6 cards = classic
        request.vars['cuID'] is expected to be the customers_id
    '''
    vars = request.vars

    query = (db.school_classcards.Archived == False)
    count_cards = db(query).count()

    if count_cards < 9:
        redirect(URL('classcard_add_modern', vars=vars))
    else:
        redirect(URL('classcard_add_classic', vars=vars))


def classcard_add_check_trialcard(cuID):
    '''
        Check whether a customer has already had a trialcard
    '''
    tc_query = (db.school_classcards.id ==
                db.customers_classcards.school_classcards_id) & \
               (db.school_classcards.Trialcard == True) & \
               (db.customers_classcards.auth_customer_id == cuID)
    tc_count = db(tc_query).count()

    return tc_count


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'customers_classcards'))
def classcard_add_modern():
    '''
        Add a new classcard for a customer in more graphic way than
        a drop down menu
        request.vars['cuID'] is expected to be the customers_id
    '''
    customers_id   = request.vars['cuID']
    clsID          = request.vars['clsID'] # for redirect to classes attendance_list_classcards
    date_formatted = request.vars['date'] # for redirect to classes attendance_list_classcards
    customer = Customer(customers_id)
    response.title = customer.get_name()
    response.subtitle = T("New Class card")
    response.view = 'general/tabs_menu.html'

    cuID = request.vars['cuID']
    return_url = classcards_get_return_url(customers_id, clsID, date_formatted)

    query = (db.school_classcards.Archived == False)

    if classcard_add_check_trialcard(customers_id):
        query &= (db.school_classcards.Trialcard == False)

    rows = db(query).select(db.school_classcards.ALL,
                            orderby=db.school_classcards.Trialcard|\
                                    db.school_classcards.Name)

    # check for no class cards
    query = (db.school_classcards.Archived == False)
    count_cards = db(query).count()


    back = DIV(os_gui.get_button('back_bs', return_url), BR(), BR(), _class='col-md-12')
    if count_cards < 1:
        content = DIV(back, T("No class cards found, please add one under school."), _class='row')
    else:
        content = DIV(back, _class='row')
    # populate as usual
    modals = DIV()
    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]
        card_name = max_string_length(row.Name, 33)
        validity = classcard_get_validity(row)

        if clsID:
            vars = {'cuID' : cuID,
                    'scdID': row.id,
                    'clsID': clsID,
                    'date' : date_formatted}
        else:
            vars = {'cuID' : cuID,
                    'scdID': row.id}

        form_id = 'form_' + unicode(row.id)

        modal_content = LOAD('customers', 'classcard_add_modern_add_card.load',
                             ajax_trap=True,
                             vars=vars)
        result =  os_gui.get_modal(button_text=T('This one'),
                                   button_class='btn-block btn-link',
                                   modal_title=card_name,
                                   modal_content=modal_content,
                                   modal_footer_content=os_gui.get_submit_button(form_id),
                                   modal_class='modal_card_' + unicode(row.id))
        modals.append(result['modal'])

        card_content = DIV(
            TABLE(TR(TD(T("Validity"), TD(validity))),
                  TR(TD(T("Classes"), TD(repr_row.Classes))),
                  TR(TD(T("Price"), TD(repr_row.Price))),
                  _class='table'),
            result['button']
            )


        if row.Trialcard:
            panel_class = 'panel-success'
        else:
            panel_class = 'panel-primary'
        card = DIV(os_gui.get_panel_table(card_name, card_content, panel_class),
                   _class='col-md-3')

        content.append(card)


        if i == 4:
            content.append(BR())

    content.append(modals)

    back = edit_get_back()

    menu = customers_get_menu(customers_id, 'classcards')

    return dict(content=content, back=back, menu=menu)


def classcard_get_validity(row):
    '''
        takes a db.school_classcards() row as argument
    '''
    validity = SPAN(unicode(row.Validity), ' ')

    validity_in = represent_validity_units(row.ValidityUnit, row)
    if row.Validity == 1:  # Cut the last 's"
        validity_in = validity_in[:-1]

    validity.append(validity_in)

    return validity


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'customers_classcards'))
def classcard_add_modern_add_card():
    """
        This function shows an add page for a classcard
        request.vars['cuID'] is expected to be the customers_id
        request.vars['scdID'] is expected to be the school_classcards.id
    """
    cuID = request.vars['cuID']
    scdID = request.vars['scdID']

    response.title = T("New Class card")
    customer = Customer(cuID)
    response.subtitle = customer.get_name()

    return_url = URL('classcard_add_modern_add_card_redirect_classcards',
                     vars=request.vars, extension='')

    classcard_add_check_trialcard_set_query(cuID)
    db.customers_classcards.school_classcards_id.default = scdID
    db.customers_classcards.school_classcards_id.readable = False
    db.customers_classcards.school_classcards_id.writable = False
    db.customers_classcards.Enddate.readable = False
    db.customers_classcards.Enddate.writable = False

    db.customers_classcards.auth_customer_id.default = cuID

    functions_onadd = classcard_add_get_functions()
    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Added card")
    crud.settings.create_next = return_url
    crud.settings.create_onaccept = functions_onadd
    form = crud.create(db.customers_classcards)

    form_id = "form_" + unicode(scdID)
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    return dict(content=form)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'customers_classcards'))
def classcard_add_modern_add_card_redirect_classcards():
    '''
        Redirect back to classcards list after adding a new card
        We need this extra function, because otherwise we're stuck in the modal
        This way we can call redirect() with client_side=True
    '''
    session.flash = T("Added card")
    cuID  = request.vars['cuID']
    clsID = request.vars['clsID']
    date_formatted  = request.vars['date']

    redirect_url = classcards_get_return_url(cuID, clsID, date_formatted)
    redirect(redirect_url, client_side=True)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'customers_classcards'))
def classcard_add_classic():
    """
        This function shows an add page for a classcard
        request.vars['cuID'] is expected to be the customers_id
    """
    customers_id   = request.vars['cuID']
    clsID          = request.vars['clsID'] # for redirect to classes attendance_list_classcards
    date_formatted = request.vars['date'] # for redirect to classes attendance_list_classcards
    customer = Customer(customers_id)
    response.title = customer.get_name()
    response.subtitle = T("New Class card")
    response.view = 'general/tabs_menu.html'

    db.customers_classcards.auth_customer_id.default = customers_id

    classcard_add_check_trialcard_set_query(customers_id)

    return_url = classcards_get_return_url(customers_id, clsID, date_formatted)

    functions_onadd = classcard_add_get_functions()

    db.customers_classcards.Enddate.readable = False
    db.customers_classcards.Enddate.writable = False

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Added card")
    crud.settings.create_next = return_url
    crud.settings.create_onaccept = functions_onadd
    form = crud.create(db.customers_classcards)

    form_element = form.element('form')
    form['_id'] = 'MainForm'

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = "MainForm"

    submit = form.element('input[type=submit]')

    back = os_gui.get_button('back_bs', return_url)

    content = DIV(back, form)

    back = edit_get_back()

    menu = customers_get_menu(customers_id, 'classcards')

    return dict(content=content, back=back, menu=menu, save=submit)



def classcard_add_check_trialcard_set_query(customers_id):
    '''
        Don't allow adding trialcard when one has already been added
    '''
    if classcard_add_check_trialcard(customers_id):
        scd_query = (db.school_classcards.Archived == False) & \
                    (db.school_classcards.Trialcard == False)
        db.customers_classcards.school_classcards_id.requires=\
            IS_IN_DB(db(scd_query),
                     'school_classcards.id',
                     '%(Name)s',
                     zero=(T('Please select...')))


def classcard_add_get_functions(var=None):
    '''
        Functions to execute after adding a classcard
    '''
    functions = [ classcard_add_set_enddate,
                  classcard_add_create_invoice,
                  classcards_clear_cache ]

    return functions


def classcard_add_set_enddate(form):
    '''
        Calculate and set enddate when adding a classcard
    '''
    # get info
    if 'school_classcards_id' in form.vars:
        # used for classic classcards
        scdID = form.vars.school_classcards_id
    else:
        # used for modern classcards
        scdID = db.customers_classcards.school_classcards_id.default

    scd = SchoolClasscard(scdID)
    enddate = scd.sell_to_customer_get_enddate(form.vars.Startdate)

    # set enddate
    row = db.customers_classcards(form.vars.id)
    row.Enddate = enddate
    row.update_record()


def classcard_add_create_invoice(form):
    '''
        Add an invoice after adding a classcard
    '''
    ccdID   = form.vars.id
    scdID   = form.vars.school_classcards_id

    scd = SchoolClasscard(scdID)
    scd.sell_to_customer_create_invoice(ccdID)


@auth.requires_login()
def classcard_edit():
    """
        This function shows an edit page for a classcard
        request.vars['cuID'] is expected to be the customers_id
        request.vars['ccdID'] is expected to be the classcardID
    """
    customers_id = request.vars['cuID']
    classcardID = request.vars['ccdID']
    response.title = T("Edit Class card") + " " + unicode(classcardID)
    customer = Customer(customers_id)
    classcard = Classcard(classcardID)
    response.subtitle = customer.get_name()
    response.view = 'general/tabs_menu.html'

    db.customers_classcards.school_classcards_id.writable = False
    db.customers_classcards.school_classcards_id.readable = False

    # permission check to prevent editing of enddate
    permission = auth.has_membership(group_id='Admins') or \
                 auth.has_permission('update', 'customers_classcards_enddate')
    if not permission:
        db.customers_classcards.Enddate.readable = False
        db.customers_classcards.Enddate.writable = False

    return_url = classcards_get_return_url(customers_id)

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Updated class card")
    crud.settings.update_onaccept = [ classcard_edit_update_classes_taken, classcards_clear_cache ]
    crud.settings.update_next = return_url
    crud.settings.update_deletable = False
    form = crud.update(db.customers_classcards, classcardID)

    form_element = form.element('form')
    form['_id'] = 'MainForm'

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = "MainForm"

    submit = form.element('input[type=submit]')

    back = os_gui.get_button('back_bs', return_url)

    content = DIV(back, form)

    back = edit_get_back()

    menu = customers_get_menu(customers_id, 'classcards')

    return dict(content=content, back=back, menu=menu, save=submit)


def classcard_edit_update_classes_taken(form):
    '''
        Updates number of classes taken when saving a classcard
    '''
    ccdh = ClasscardsHelper()
    ccdh.set_classes_taken(form.vars.id)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'customers_classcards'))
def classcard_classes():
    response.title = T("Class cards")
    ccdID = request.vars['ccdID']
    back = request.vars['back']
    response.subtitle = T("Classes taken using card") + " " + ccdID
    response.view = 'general/only_content.html'
    row = db.customers_classcards(ccdID)
    customers_id = row.auth_customer_id

    classcard = Classcard(ccdID)
    rows = classcard.get_rows_classes_taken()

    table = TABLE(TR(TH(T("Class date")),
                     TH(T("Location")),
                     TH(T("Type")),
                     TH(T("Start")),
                     _class='header'),
                  _class='table')
    for i, row in enumerate(rows):
        repr_row = list(rows[i:i+1].render())[0]

        clsID = row.classes.id
        date = row.classes_attendance.ClassDate
        teachers = class_get_teachers(clsID, date)

        table.append(TR(TD(repr_row.classes_attendance.ClassDate),
                        TD(repr_row.classes.school_locations_id),
                        TD(repr_row.classes.school_classtypes_id),
                        TD(repr_row.classes.Starttime)))

    back = os_gui.get_button('back',
                             URL('classcards', vars={'cuID':customers_id}))

    return dict(content=table, back=back)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'classes_reservation'))
def classes_reservations():
    '''
        Show reservations for a customer
    '''
    response.view = 'customers/edit_general.html'

    cuID = request.vars['cuID']
    session.customers_classes_reservation_add_vars = {}
    session.customers_classes_reservation_add_vars['cuID'] = cuID
    session.classes_reserve_back = 'customers_reservations'
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Classes")

    submenu = classes_get_submenu(request.function, cuID)
    content = DIV(submenu)

    add = os_gui.get_button('add',
                            URL('classes_reservation_add', vars={'cuID':cuID}),
                            btn_size='btn-sm',
                            _class='pull-right')

    if 'filter' in request.vars:
        session.customers_reservations_filter = request.vars['filter']
    elif session.customers_reservations_filter:
        pass
    else:
        session.customers_reservations_filter = 'recurring'

    # buttons = [ [ 'recurring', T('Enrollments') ],
    #             [ 'single', T('Drop in') ],
    #             [ 'trial', T('Trial') ] ]
    # filter_form = os_gui.get_radio_buttons_form(
    #     session.customers_reservations_filter,
    #     buttons)

    content.append(BR())
    #tools = DIV(filter_form, _class='pull-right')
    #content.append(tools)
    #content.append(BR())
    #content.append(BR())

    # list of reservations
    db.classes_reservation.id.readable = False
    db.classes_reservation.auth_customer_id.readable = False
    links = ''
    headers = {'classes.Starttime' : T('Time'),
               'classes_reservation.Startdate' : T('Class date')}
    fields = [
          db.classes.Week_day,
          db.classes.Starttime,
          db.classes.school_locations_id,
          db.classes.school_classtypes_id,
          db.classes_reservation.Startdate,
    ]

    query = (db.classes_reservation.auth_customer_id == cuID)
    if session.customers_reservations_filter == 'recurring':
        query &= (db.classes_reservation.SingleClass == False)
        query &= (db.classes_reservation.TrialClass == False)

        fields.append(db.classes_reservation.Enddate)
        headers.pop('classes_reservation.Startdate')

        links = [ lambda row: os_gui.get_button(
                        'edit', URL('classes', 'reservation_edit',
                                    vars={'crID':row.classes_reservation.id})) ]
    elif session.customers_reservations_filter == 'single':
        query &= (db.classes_reservation.SingleClass == True)
        query &= (db.classes_reservation.TrialClass == False)
    elif session.customers_reservations_filter == 'trial':
        query &= (db.classes_reservation.SingleClass == True)
        query &= (db.classes_reservation.TrialClass == True)


    maxtextlengths = {'classes_reservation.classes_id' : 50}
    left = db.classes.on(db.classes_reservation.classes_id == db.classes.id)

    delete_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('delete', 'classes_reservation')

    grid = SQLFORM.grid(query,
                        fields=fields,
                        headers=headers,
                        links=links,
                        details=False,
                        searchable=False,
                        deletable=delete_permission,
                        csv=False,
                        create=False,
                        editable=False,
                        maxtextlengths=maxtextlengths,
                        left=left,
                        orderby=~db.classes_reservation.Startdate,
                        field_id=db.classes_reservation.id,
                        ui = grid_ui)
    grid.element('.web2py_counter', replace=None) # remove the counter
    grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    content.append(grid)

    # Back button
    back = edit_get_back()
    menu = customers_get_menu(cuID, 'classes')

    return dict(content=content,
                back=back,
                add=add,
                menu=menu,
                left_sidebar_enabled=True)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'classes_reservation'))
def classes_reservation_add():
    '''
        Add a new reservation for a customer on the selected date
        request.vars['cuID'] is expected to be db.customers.id
    '''
    response.view = 'general/only_content.html'

    cuID = request.vars['cuID']
    session.customers_classes_reservation_add_vars = {}
    session.customers_classes_reservation_add_vars['cuID'] = cuID
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Enroll in a class")

    session.classes_reserve_back = 'customers_reservations'

    if 'date' in request.vars:
        # response.subtitle = SPAN(T('for'), ' ',
        #                          customer.get_name(), ' ',
        #                          request.vars['date'])
        default_date = datestr_to_python(DATE_FORMAT, request.vars['date'])
    else:
        default_date = datetime.date.today()

    session.customers_classes_reservation_add_vars['date'] = default_date


    result = classes_add_get_form_date(cuID, default_date)
    form = result['form']
    form_date = result['form_styled']


    db.classes.id.readable = False

    # list of classes
    grid = classes_add_get_list(default_date, 'reservations')

    back = os_gui.get_button('back', URL('classes_reservations',
                                         vars={'cuID':cuID}),
                             _class='left')

    return dict(content=DIV(form_date, grid),
                back=back)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'classes_attendance'))
def classes_attendance_add():
    '''
        Add customers to attendance for a class
    '''
    response.view = 'general/only_content.html'

    cuID = request.vars['cuID']
    session.customers_classes_attendance_add_vars = {}
    session.customers_classes_attendance_add_vars['cuID'] = cuID
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Add attendance")


    if 'date' in request.vars:
        # response.subtitle = SPAN(T('for'), ' ',
        #                          customer.get_name(), ' ',
        #                          request.vars['date'])
        date = datestr_to_python(DATE_FORMAT, request.vars['date'])
    else:
        date = datetime.date.today()

    session.customers_classes_attendance_add_vars['date'] = date


    result = classes_add_get_form_date(cuID, date)
    form = result['form']
    form_date = result['form_styled']


    db.classes.id.readable = False

    # list of classes
    grid = classes_add_get_list(date, 'attendance', cuID)

    back = os_gui.get_button('back', URL('classes_attendance',
                                         vars={'cuID':cuID}),
                             _class='left')

    return dict(content=DIV(form_date, grid),
                back=back)


@auth.requires(auth.has_membership(group_id='Admins') or
               auth.has_permission('create', 'classes_attendance'))
def classes_attendance_add_booking_options():
    '''
        Page to show listing of booking options for customer
    '''
    cuID = request.vars['cuID']
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    date = datestr_to_python(DATE_FORMAT, date_formatted)

    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Add attendance")
    response.view = 'general/only_content.html'

    return_url = URL('customers', 'classes_attendance_add', vars={'cuID':cuID})

    cls = Class(clsID, date)

    ah = AttendanceHelper()
    options = ah.get_customer_class_booking_options(clsID,
                                                    date,
                                                    customer,
                                                    trial=True,
                                                    complementary=True,
                                                    list_type='attendance',
                                                    controller='classes')
    cancel = os_gui.get_button('noicon',
                               return_url,
                               title=T('Cancel'),
                               btn_size='')

    content = DIV(
        H4(T('Booking options for class'), ' ', cls.get_name(), _class='center'), BR(),
        options,
        DIV(BR(), cancel, _class="col-md-12 center"),
        _class="row"
    )

    back = os_gui.get_button('back', return_url)

    return dict(content=content,
                back=back)



def classes_add_get_form_date(cuID, date):
    '''
        Get date form
    '''
    form = SQLFORM.factory(
        Field('date', 'date',
            requires=IS_DATE_IN_RANGE(format=DATE_FORMAT,
                          minimum=datetime.date(1900,1,1),
                          maximum=datetime.date(2999,1,1)),
            default=date,
            widget=os_datepicker_widget),
        separator = '',
        submit_button = T('Go'))

    input_date = form.element('#no_table_date')
    #input_date.attributes['_onchange'] = "this.form.submit();"

    submit = form.element('input[type=submit]')

    delta = datetime.timedelta(days=1)
    date_prev = (date - delta).strftime(DATE_FORMAT)
    date_next = (date + delta).strftime(DATE_FORMAT)

    url_prev = URL(vars={'cuID': cuID,
                         'date': date_prev})
    url_next = URL(vars={'cuID': cuID,
                         'date': date_next})

    previous = A(I(_class='fa fa-angle-left'),
                 _href=url_prev,
                 _class='btn btn-default')
    nxt = A(I(_class='fa fa-angle-right'),
            _href=url_next,
            _class='btn btn-default')

    chooser = DIV(previous, nxt, _class='btn-group pull-right')

    form_styled = DIV(form.custom.begin,
                      DIV(B('Show classes on '),
                          form.custom.widget.date,
                          _class='col-md-3'),
                      DIV(BR(),
                          form.custom.submit,
                          chooser,
                          _class='col-md-9 no-padding-left'),
                      form.custom.end,
                      _class='row')


    return {'form':form,
            'form_styled':form_styled}


def classes_add_get_list(date, list_type, cuID=None):
    '''
        Get list of classes for a date
    '''
    if list_type == 'attendance':
        session.classes_attendance_signin_back = 'cu_classes_attendance'
        ah = AttendanceHelper()
        #links = [ lambda row: ah.get_signin_buttons(row.classes.id, date, cuID) ]

    if session.classes_schedule_sort == 'location':
        orderby = db.school_locations.Name | db.classes.Starttime
    elif session.classes_schedule_sort == 'starttime':
        orderby = db.classes.Starttime | db.school_locations.Name
    else:
        orderby = db.school_locations.Name | db.classes.Starttime

    cs = ClassSchedule(date, sorting=orderby)
    classes = cs.get_day_list()

    header = THEAD(TR(TH(T('Time')),
                      TH(T('Location')),
                      TH(T('Class')),
                      TH(),
                      TH() # buttons
                      ))
    table = TABLE(header, _class='table table-striped table-hover')
    for c in classes:
        status = classes_add_get_list_get_cancelled_holiday(c)
        buttons = ''

        if list_type == 'reservations':
            buttons = classes_reservation_add_get_button(c['ClassesID'])
        if list_type == 'attendance' and status == '':
            buttons = os_gui.get_button('noicon',
                                        URL('customers', 'classes_attendance_add_booking_options',
                                            vars={'cuID':cuID,
                                                  'clsID':c['ClassesID'],
                                                  'date':date.strftime(DATE_FORMAT)}),
                                        title='Check in',
                                        _class='pull-right')

        tr = TR(
            TD(c['Starttime'], ' - ', c['Endtime']),
            TD(c['Location']),
            TD(c['ClassType']),
            TD(status),
            TD(buttons)
        )

        table.append(tr)

    return table


def classes_add_get_list_get_cancelled_holiday(c):
    '''
        Returns class or holiday description when a class is cancelled
        :param: class from ClassSchedule.get_day_list()
    '''
    status = ''

    if c['Cancelled']:
        status = SPAN(T('Cancelled'), ' ', SPAN(c['CancelledDescription'], _class='grey'))

    if c['Holiday']:
        status = SPAN(T('Holiday'), ' ', SPAN(c[''], _class='grey'))

    return status


def classes_reservation_add_get_button(clsID):
    '''
        Returns add button for a customer to add a reservation
    '''
    date = session.customers_classes_reservation_add_vars['date']
    date_formatted = date.strftime(DATE_FORMAT)
    cuID = session.customers_classes_reservation_add_vars['cuID']
    customer = Customer(cuID)

    add = os_gui.get_button('add', URL('classes', 'reservation_add',
                                       vars={'cuID':cuID, 'clsID':clsID, 'date':date_formatted}),
                            btn_size='btn-sm',
                            _class="pull-right")

    return add


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'classes_waitinglist'))
def classes_waitinglist():
    '''
        Show waitinglist for a customer
    '''
    response.view = 'customers/edit_general.html'

    cuID = request.vars['cuID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Classes")

    submenu = classes_get_submenu(request.function, cuID)
    content = DIV(submenu)

    delete_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('delete', 'classes_waitinglist')

    # fill content area
    query = (db.classes_waitinglist.auth_customer_id == cuID)
    fields = [ db.classes.Week_day,
               db.classes.school_locations_id,
               db.classes.school_classtypes_id,
               db.classes.Starttime ]
    wai_grid = SQLFORM.grid(query, fields,
                            create=False,
                            details=False,
                            editable=False,
                            csv=False,
                            searchable=False,
                            deletable=delete_permission,
                            user_signature=False,
                            field_id=db.classes_waitinglist.id,
                            left=db.classes.on(
                                    db.classes_waitinglist.classes_id==
                                    db.classes.id),
                            ui = grid_ui)
    wai_grid.element('.web2py_counter', replace=None) # remove the counter
    # remove text from delete button
    wai_grid.elements('span[title=Delete]', replace=None)
    wai_title = H3(T("Waitinglist"))

    content.append(wai_grid)

    # Back button
    back = edit_get_back()
    menu = customers_get_menu(cuID, 'classes')

    return dict(content=content,
                back=back,
                menu=menu,
                left_sidebar_enabled=True)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'classes_attendance'))
def classes_attendance():
    '''
        Show waitinglist for a customer
    '''
    response.view = 'customers/edit_general.html'

    session.classes_attendance_remove_back = 'customers'
    session.invoices_payment_add_back = 'customers_classes_attendance'

    cuID = request.vars['cuID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Classes")

    if 'all' in request.vars:
        limit = False
        limit_by = False
        link_all = ''
    else:
        limit = 25
        limit_by = limit + 1
        link_all = A(T('Show all'), _href=URL(vars={'cuID':cuID,
                                                    'all': True}))

    submenu = classes_get_submenu(request.function, cuID)
    content = DIV(submenu)
    add = os_gui.get_button('add',
                            URL('classes_attendance_add', vars={'cuID':cuID}),
                            btn_size='btn-sm',
                            _class='pull-right')
    content.append(BR())

    header = THEAD(TR(TH(T('Date')),
                      TH(T('Time')),
                      TH(T('Class')),
                      TH(T('Location')),
                      TH(T('Used')),
                      TH(T('Status')),
                      TH(),
                      TH())) # Actions

    table = TABLE(header, _class='table table-striped table-hover')

    edit_permission = (auth.has_membership(group_id='Admins') or
                       auth.has_permission('update', 'classes_attendance'))
    delete_permission = (auth.has_membership(group_id='Admins') or
                         auth.has_permission('delete', 'classes_attendance'))


    rows = customer.get_classes_attendance_rows(limit_by)


    for i, row in enumerate(rows):
        if limit:
            if i + 1 > limit:
                break

        repr_row = list(rows[i:i + 1].render())[0]

        cancel = ''
        if edit_permission and not row.classes_attendance.BookingStatus == 'cancelled':
            cancel = classes_attendance_get_link_cancel(row)

        remove = ''
        if delete_permission:
            remove = classes_attendance_get_link_delete(row)

        ##
        # Attendance type labels
        ##
        ct = ''
        if row.classes_attendance.AttendanceType == 1:
            # trial class
            ct = os_gui.get_label('success', repr_row.classes_attendance.AttendanceType)
        elif row.classes_attendance.AttendanceType == 2:
            # drop in class
            ct = os_gui.get_label('primary', repr_row.classes_attendance.AttendanceType)
        elif row.classes_attendance.AttendanceType == 4:
            # Complementary class
            ct = os_gui.get_label('default', repr_row.classes_attendance.AttendanceType)


        ##
        # Invoice
        ##
        invoice = ''
        if row.invoices.id:
            ih = InvoicesHelper()
            invoice = ih.represent_invoice_for_list(
                row.invoices.id,
                repr_row.invoices.InvoiceID,
                repr_row.invoices.Status,
                row.invoices.Status,
                row.invoices.payment_methods_id
            )


        att_type = ''
        if row.classes_attendance.customers_subscriptions_id:
            att_type = repr_row.classes_attendance.customers_subscriptions_id
        elif row.classes_attendance.customers_classcards_id:
            att_type = SPAN(row.school_classcards.Name,
                            _title=T('Class card') + ' ' + unicode(row.classes_attendance.customers_classcards_id))

        tr = TR(TD(repr_row.classes_attendance.ClassDate),
                TD(SPAN(repr_row.classes.Starttime, ' - ',
                        repr_row.classes.Endtime)),
                TD(repr_row.classes.school_classtypes_id),
                TD(repr_row.classes.school_locations_id),
                TD(att_type),
                TD(repr_row.classes_attendance.BookingStatus, ' ', ct),
                TD(invoice),
                TD(remove, cancel)
                )
        table.append(tr)

    content.append(table)

    # determine whether to show show all link
    if limit:
        if len(rows) <= limit:
            link_all = ''
    content.append(link_all)

    # Back button
    back = edit_get_back()
    menu = customers_get_menu(cuID, 'classes')

    return dict(content=content,
                back=back,
                add=add,
                menu=menu,
                left_sidebar_enabled=True)


def classes_attendance_get_link_cancel(row):
    '''
        Returns cancel button for classes_attendance
    '''
    button = ''

    onclick_cancel = "return confirm('" + \
     T('Do you really want to cancel this booking and refund any credits associated with it?')\
     + "');"

    button = os_gui.get_button('cancel_notext',
       URL('customers', 'classes_attendance_cancel',
           vars={'caID':row.classes_attendance.id}),
       onclick=onclick_cancel,
       _class='pull-right')

    return button


def classes_attendance_get_link_delete(row):
    '''
        Checks delete permissions and returns button if granted
    '''
    button = ''

    onclick_delete = "return confirm('" + \
     T('Do you really want to remove this class from the attendance list?')\
     + "');"

    date = row.classes_attendance.ClassDate.strftime(DATE_FORMAT)
    cuID = row.classes_attendance.auth_customer_id

    button = os_gui.get_button('delete_notext',
       URL('classes', 'attendance_remove',
           vars={'clattID':row.classes_attendance.id}),
       onclick=onclick_delete,
       _class='pull-right')

    return button


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('update', 'classes_attendance'))
def classes_attendance_cancel():
    '''
        Actually cancel a booking
        request.vars['caID'] is expected to be from db.classes_attendance.id
    '''
    caID = request.vars['caID']
    clatt = ClassAttendance(caID)

    clatt.set_status_cancelled(force=True)

    redirect(URL('customers', 'classes_attendance', vars={'cuID':clatt.row.auth_customer_id}))


def classes_get_submenu(page, cuID):
    """
        This function returns a submenu for the classes edit pages
    """
    vars = {'cuID':cuID}
    pages = [['classes_attendance', T('Attendance'),
              URL('classes_attendance', vars=vars)],
             ['classes_reservations', T('Enrollments'),
               URL('classes_reservations', vars=vars)],
             ['classes_waitinglist', T('Waitinglist'),
               URL('classes_waitinglist', vars=vars)]]

    horizontal = True

    return os_gui.get_submenu(pages, page, horizontal=horizontal, htype='tabs')


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'classes_reservation'))
def classes_edit():
    customers_id = request.vars['cuID']
    session.classes_reserve_back = 'customers'
    response.view = 'general/one_grid_with_sidebar.html'
    response.title = T("Edit class reservations")
    row = db.customers[customers_id]
    response.subtitle = row.display_name

    query = (db.classes)
    links = [classes_check_reservation]
    fields = [ db.classes.Week_day,
               db.classes.school_locations_id,
               db.classes.school_classtypes_id,
               db.classes.Starttime ]
    grid = SQLFORM.grid(query,
                        fields,
                        create=False,
                        details=False,
                        deletable=False,
                        editable=False,
                        searchable=False,
                        csv=False,
                        links=links,
                        paginate=50,
                        orderby=db.classes.Week_day|\
                                db.classes.school_locations_id|\
                                db.classes.Starttime,
                        ui=grid_ui)
    grid.element('.web2py_counter', replace=None) # remove the counter

    back = os_gui.get_button('back', URL('classes', vars={'cuID':customers_id}))

    return dict(grid=grid, back=back)


def subscription_credits_clear_cache(form):
    '''
        Clear the subscriptions cache for customer
    '''
    cscID = form.vars.id
    csc = db.customers_subscriptions_credits(cscID)
    cs = db.customers_subscriptions(csc.customers_subscriptions_id)
    cuID = cs.auth_customer_id

    cache_clear_customers_subscriptions(cuID)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_subscriptions_credits'))
def subscription_credits():
    '''
        This function shows a page listing the credits total for a subscription
        request.vars['csID'] is expected to be db.customers_subscriptions.id
    '''
    response.view = 'general/tabs_menu.html'
    cuID = request.vars['cuID']
    csID = request.vars['csID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = subscription_edit_get_subtitle(csID)

    cs = CustomerSubscription(csID)

    total = H4(T('Balance:'), ' ', cs.get_credits_balance(), _class='pull-right')
    mutations = cs.get_credits_mutations_rows(formatted=True,
                                              editable=True,
                                              deletable=True,
                                              delete_controller='customers',
                                              delete_function='subscription_credits_delete')

    add = os_gui.get_button('add',
                            URL('subscription_credits_add', vars={'csID':csID,
                                                                  'cuID':cuID}),
                            btn_size='btn-sm')

    back = subscription_edit_get_back(cuID)
    menu = subscription_edit_get_menu(cuID, csID, request.function)

    return dict(content=DIV(total, mutations), menu=menu, back=back, add=add)


def subscription_credits_get_return_url(cuID, csID):
    '''
        Return url for subscription credits
    '''
    return URL('customers', 'subscription_credits', vars={'cuID':cuID,
                                                          'csID':csID})


@auth.requires_login()
def subscription_credits_add():
    """
        This function shows an add page for subscription credits
        request.vars['cuID'] is expected to be auth.user.id
        request.vars['csID'] is expected to be customers_subscriptions_id
    """
    response.view = 'general/tabs_menu.html'
    cuID = request.vars['cuID']
    csID = request.vars['csID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = subscription_edit_get_subtitle(csID)

    db.customers_subscriptions_credits.customers_subscriptions_id.default = csID

    return_url = subscription_credits_get_return_url(cuID, csID)

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Added subscription credit(s)")
    crud.settings.create_next = return_url
    crud.settings.create_onaccept = [subscription_credits_clear_cache]
    form = crud.create(db.customers_subscriptions_credits)

    element_form = form.element('form')
    element_form['_id'] = "MainForm"

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = "MainForm"

    submit = form.element('input[type=submit]')


    subscr_back = os_gui.get_button('back_bs', URL('subscription_credits', vars={'cuID':cuID,
                                                                                 'csID':csID}))
    content = DIV(subscr_back, form)

    back = os_gui.get_button("back", return_url, _class='')

    menu = customers_get_menu(cuID, 'subscriptions_credits')

    return dict(content=content, back=back, menu=menu, save=submit)


@auth.requires_login()
def subscription_credits_edit():
    """
        This function shows an edit page for  subscription credits
        request.vars['cuID'] is expected to be auth.user.id
        request.vars['csID'] is expected to be customers_subscriptions_id
    """
    response.view = 'general/tabs_menu.html'
    cuID = request.vars['cuID']
    csID = request.vars['csID']
    cscID = request.vars['cscID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = subscription_edit_get_subtitle(csID)

    db.customers_subscriptions_credits.customers_subscriptions_id.default = csID

    return_url = subscription_credits_get_return_url(cuID, csID)

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.update_next = return_url
    crud.settings.update_onaccept = [subscription_credits_clear_cache]
    form = crud.update(db.customers_subscriptions_credits, cscID)

    element_form = form.element('form')
    element_form['_id'] = "MainForm"

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = "MainForm"

    submit = form.element('input[type=submit]')


    subscr_back = os_gui.get_button('back_bs', URL('subscription_credits', vars={'cuID':cuID,
                                                                                 'csID':csID}))
    content = DIV(subscr_back, form)

    back = os_gui.get_button("back", return_url, _class='')

    menu = customers_get_menu(cuID, 'subscriptions_credits')

    return dict(content=content, back=back, menu=menu, save=submit)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('delete', 'customers_subscriptions_credits'))
def subscription_credits_delete():
    '''
        Delete subscription credits
    '''
    cuID = request.vars['cuID']
    csID = request.vars['csID']
    cscID = request.vars['cscID']

    query = (db.customers_subscriptions_credits.id == cscID)
    db(query).delete()

    cache_clear_customers_subscriptions(cuID)

    redirect(URL('subscription_credits', vars={'cuID':cuID, 'csID':csID}))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'customers_subscriptions'))
def subscription_pauses():
    '''
        This function shows a page which lists all pauses for a subscription
        request.vars['csID'] is expected to be the subscription ID
    '''
    response.view = 'general/tabs_menu.html'
    cuID = request.vars['cuID']
    csID = request.vars['csID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = subscription_edit_get_subtitle(csID)

    row = db.customers_subscriptions(csID)
    db.customers_subscriptions_paused.id.readable = False

    query = (db.customers_subscriptions_paused.customers_subscriptions_id == csID)
    if db(query).count() == 0:
        grid = DIV(BR(), T("This subscription hasn't been paused before."))
    else:
        grid = SQLFORM.grid(query,
            create=False,
            details=False,
            editable=False,
            searchable=False,
            csv=False,
            paginate=50,
            orderby=db.customers_subscriptions_paused.Startdate,
            field_id=db.customers_subscriptions_paused.id,
            ui = grid_ui)
        grid.element('.web2py_counter', replace=None) # remove the counter
        grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    add = os_gui.get_button('add', URL('subscription_pause_add', vars={'csID':csID}), btn_size='btn-sm')

    back = subscription_edit_get_back(cuID)
    menu = subscription_edit_get_menu(cuID, csID, request.function)

    return dict(content=grid, menu=menu, back=back, add=add)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'customers_subscriptions'))
def subscription_pause_add():
    '''
        This function shows a page to allow a user to pause a subscription for multiple months
        request.vars['csID'] is expected to be the customers_subscriptions.id
    '''
    response.view = 'general/tabs_menu.html'
    csID = request.vars['csID']
    cs   = CustomerSubscription(csID)
    cuID = cs.auth_customer_id
    customer = Customer(cuID)

    response.title = customer.get_name()
    response.subtitle = SPAN(T("Pause subscription"), ' ',
                             cs.name)

    today = datetime.date.today()
    return_url = URL('subscription_pauses', vars={'cuID':cuID,
                                                  'csID':csID})

    months = get_months_list()

    form = SQLFORM.factory(
        Field('from_month', 'integer',
            requires=IS_IN_SET(months),
            default=today.month),
        Field('from_year', 'integer',
            default=today.year),
        Field('until_month', 'integer',
            requires=IS_IN_SET(months),
            default=today.month),
        Field('until_year', 'integer',
            default=today.year),
        Field('description'),
        separator = '',
        submit_button = T("Save")
        )

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    # change type to number from the default text
    form.element('#no_table_from_year').attributes['_type'] = 'number'
    form.element('#no_table_from_year').attributes['_class'] += ' os-input_year'
    form.element('#no_table_from_month').attributes['_class'] += ' os-input_month'
    form.element('#no_table_until_year').attributes['_type'] = 'number'
    form.element('#no_table_until_year').attributes['_class'] += ' os-input_year'
    form.element('#no_table_until_month').attributes['_class'] += ' os-input_month'
    form.element('#no_table_description').attributes['_placeholder'] = T("Description")


    if form.process().accepted:
        start = datetime.date(int(form.vars.from_year), int(form.vars.from_month), 1)
        end = get_last_day_month(datetime.date(int(form.vars.until_year), int(form.vars.until_month), 1))
        description = form.vars.description
        db.customers_subscriptions_paused.insert(customers_subscriptions_id = csID,
                                                 Startdate = start,
                                                 Enddate = end,
                                                 Description = description)
        redirect(return_url)


    back = os_gui.get_button('back_bs', return_url)
    content = DIV(
        DIV(back,
            XML('<form action="#" enctype="multipart/form-data" id="MainForm" method="post">'),
            TABLE(
                TR(TD(LABEL(T("From the start of "))),
                   TD(form.custom.widget.from_month),
                   TD(form.custom.widget.from_year)),
                TR(TD(LABEL(T("Until the end of "))),
                   TD(form.custom.widget.until_month),
                   TD(form.custom.widget.until_year)),
                TR(TD(),
                   TD(form.custom.widget.description, _colspan='2')),
                _id='customer_subscription_pause_add'),
            BR(),
            form.custom.end,
            _class="col-md-12"),
        _class='row')


    back = subscription_edit_get_back(cuID)
    menu = subscription_edit_get_menu(cuID, csID, 'subscription_pauses')

    return dict(content=content, menu=menu, back=back, save=submit)


def subscriptions_get_return_url(customers_id):
    '''
        Returns return URL for subscriptions
    '''
    return URL('subscriptions', vars={'cuID':customers_id})


def subscriptions_clear_cache(form):
    '''
        Clear the subscriptions cache for customer 
    '''
    csID = form.vars.id
    cs = db.customers_subscriptions(csID)
    cuID = cs.auth_customer_id

    cache_clear_customers_subscriptions(cuID)


@auth.requires_login()
def subscription_add():
    """
        This function shows an add page for a subscription
        request.args[0] is expected to be the customers_id
    """
    customers_id = request.vars['cuID']
    customer = Customer(customers_id)
    response.view = 'general/tabs_menu.html'
    response.title = customer.get_name()
    response.subtitle = T("New subscription")

    db.customers_subscriptions.auth_customer_id.default = customers_id

    return_url = subscriptions_get_return_url(customers_id)

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Added subscription")
    crud.settings.create_next = return_url
    crud.settings.create_onaccept = [subscriptions_clear_cache, subscription_add_add_credits]
    form = crud.create(db.customers_subscriptions)

    element_form = form.element('form')
    element_form['_id'] = "MainForm"

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = "MainForm"

    submit = form.element('input[type=submit]')


    subscr_back = os_gui.get_button('back_bs', URL('subscriptions', vars={'cuID':customers_id}))
    content = DIV(subscr_back, form)

    back = os_gui.get_button("back", return_url, _class='')

    menu = customers_get_menu(customers_id, 'subscriptions')

    return dict(content=content, back=back, menu=menu, save=submit)


def subscription_add_add_credits(form):
    '''
        Add credits when adding a subscription
    '''
    csID = form.vars.id
    date = form.vars.Startdate

    cs = CustomerSubscription(csID)
    cs.add_credits_month(date.year, date.month)


@auth.requires_login()
def subscription_edit():
    """
        This function shows an edit page for a subscription
        request.args[0] is expected to be the customers_id
        request.args[1] is expected to be the subscriptionID
    """
    response.view = 'general/tabs_menu.html'
    cuID = request.vars['cuID']
    csID = request.vars['csID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = subscription_edit_get_subtitle(csID)

    return_url = subscriptions_get_return_url(cuID)

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Updated subscription")
    crud.settings.update_next = return_url
    crud.settings.update_onaccept = [subscriptions_clear_cache, subscription_edit_onaccept]
    crud.settings.update_deletable = False
    form = crud.update(db.customers_subscriptions, csID)

    element_form = form.element('form')
    element_form['_id'] = "MainForm"

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = "MainForm"

    submit = form.element('input[type=submit]')


    back = subscription_edit_get_back(cuID)
    menu = subscription_edit_get_menu(cuID, csID, request.function)

    return dict(content=form,
                menu=menu,
                save=submit,
                back=back)


def subscription_edit_onaccept(form):
    '''
        :param form: CRUD form from db.customers_subscriptions
        :return: None
    '''
    if not form.vars.Enddate is None:
        enddate = form.vars.Enddate

        query = (db.classes_attendance.customers_subscriptions_id == form.vars.id) & \
                ((db.classes_attendance.ClassDate > enddate) & (db.classes_attendance.ClassDate >= TODAY_LOCAL))
        rows = db(query).select(db.classes_attendance.id)
        ids = []
        for row in rows:
            ids.append(row.id)

        query = (db.classes_attendance.id.belongs(ids))
        db(query).update(BookingStatus = 'cancelled')



@auth.requires(auth.has_membership(group_id='Admins') or
               auth.has_permission('delete', 'customers_subscriptions'))
def subscription_delete():
    '''
        Delete a subscription if no invoices are linked to it
        Otherwise display a message saying to have a look at those first
    '''
    response.view = 'general/only_content.html'
    cuID = request.vars['cuID']
    csID = request.vars['csID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    cs = CustomerSubscription(csID)
    response.subtitle = SPAN(T("Delete subscription"), ': ', cs.name)

    query = (db.invoices.customers_subscriptions_id == csID)
    invoice_count = db(query).count()

    if invoice_count:
        content = DIV(
            H3(T("Unable to delete")),
            T("One or more invoices is linked to this subscription.")
        )
    else:
        session.flash = T('Deleted subscription')
        cache_clear_customers_subscriptions(cuID)

        query = (db.customers_subscriptions.id == csID)
        db(query).delete()

        cache_clear_customers_subscriptions(cuID)

        redirect(URL('subscriptions', vars={'cuID':cuID}))


    back = subscription_edit_get_back(cuID)

    return dict(content = content,
                back=back)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'invoices'))
def subscription_invoices():
    '''
        Page to list invoices for a subscription
    '''
    if 'csID' in request.vars:
        csID = request.vars['csID']
        session.customers_subscription_invoices_csID = csID
    elif session.customers_subscription_invoices_csID:
        csID = session.customers_subscription_invoices_csID


    cuID  = request.vars['cuID']
    response.view = 'general/tabs_menu.html'

    session.invoices_edit_back = 'customers_subscription_invoices'
    session.invoices_payment_add_back = 'customers_subscription_invoices'

    # Always reset filter
    session.invoices_list_status = None

    customer = Customer(cuID)
    cs = CustomerSubscription(csID)
    response.title = customer.get_name()
    response.subtitle = SPAN(T("Edit subscription"), ' ',
                             cs.name)

    # add button
    ih = InvoicesHelper()
    form = ih.add_get_form(cuID, csID)
    result = ih.add_get_modal(form)
    add = result['button']
    modal_class = result['modal_class']

    status_filter = ih.list_get_status_filter()

    if len(form.errors):
        response.js = "show_add_modal();"

    list = ih.list_invoices(cuID=cuID, csID=csID)

    # main list
    content = DIV(DIV(status_filter,list))
                      # DIV(LOAD('invoices', 'list_invoices.load',
                      #          ajax_trap=True,
                      #          vars=request.vars,
                      #          content=os_gui.get_ajax_loader())),
                      # _class='col-md-12'),
                  #_class="row")


    menu = subscription_edit_get_menu(cuID, csID, request.function)

    back = subscription_edit_get_back(cuID)

    return dict(content=content,
                menu=menu,
                add=add,
                back=back,
                form_add=form,
                modal_class=modal_class,
                left_sidebar_enabled=True)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_subscriptions_alt_prices'))
def subscription_alt_prices():
    '''
        Specify a different price for a given month for a subscription
    '''
    csID = request.vars['csID']
    cuID  = request.vars['cuID']
    response.view = 'general/tabs_menu.html'

    customer = Customer(cuID)
    cs = CustomerSubscription(csID)
    response.title = customer.get_name()
    response.subtitle = subscription_edit_get_subtitle(csID)

    csap = db.customers_subscriptions_alt_prices
    ## main listing
    # set the default value in the grid
    links = [ lambda row: os_gui.get_button('edit',
                URL('subscription_alt_prices_edit', vars={'cuID':cuID,
                                                          'csID':csID,
                                                          'csapID':row.id})),
              lambda row: os_gui.get_button('repeat',
                                            URL('subscription_alt_price_repeat', vars={'csapID':row.id}),
                                            tooltip=T("Repeat")) ]

    query = (csap.customers_subscriptions_id == csID)

    fields = [ csap.SubscriptionMonth,
               csap.SubscriptionYear,
               csap.Amount,
               csap.Description,
               csap.Note ]

    maxtextlengths = {'customers_subscriptions_alt_prices.Description' : 36,
                      'customers_subscriptions_alt_prices.Note' : 36}
    delete_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('delete', 'customers_subscriptions_alt_prices')

    grid = SQLFORM.grid(query,
                        links=links,
                        fields=fields,
                        details=False,
                        searchable=False,
                        deletable=delete_permission,
                        csv=False,
                        create=False,
                        editable=False,
                        maxtextlengths=maxtextlengths,
                        orderby=~csap.SubscriptionYear | ~csap.SubscriptionMonth,
                        field_id=db.customers_subscriptions_alt_prices.id,
                        ui = grid_ui)
    grid.element('.web2py_counter', replace=None) # remove the counter
    grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    ## add form
    add = ''
    add_permission = auth.has_membership(group_id='Admins') or \
                     auth.has_permission('create',
                                         'customers_subscriptions_alt_prices')
    if add_permission:
        result = subscription_alt_prices_get_add_form(cuID, csID)
        form = result['form']
        submit = result['submit']
        modal_class = generate_password()
        button_text = XML(SPAN(SPAN(_class='fa fa-plus'), ' ',
                          T("Add")))
        result = os_gui.get_modal(button_text=button_text,
                                  button_title=T("Add alternative price"),
                                  modal_title=T("Add alternative price"),
                                  modal_content=form,
                                  modal_footer_content=submit,
                                  modal_class=modal_class,
                                  button_class='btn-sm')

        add = SPAN(result['button'], result['modal'])

    menu = subscription_edit_get_menu(cuID, csID, request.function)

    back = subscription_edit_get_back(cuID)

    return dict(content=grid,
                menu=menu,
                back=back,
                add=add,
                form_add=form,
                modal_class=modal_class,
                left_sidebar_enabled=True)


def subscription_alt_prices_get_add_form(cuID, csID, full_width=True):
    '''
        Returns add form for an invoice
    '''
    csap = db.customers_subscriptions_alt_prices
    csap.customers_subscriptions_id.default = csID

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Saved")
    crud.settings.create_next = URL('customers', 'subscription_alt_prices',
                                    vars={'cuID':cuID,
                                          'csID':csID})
    form = crud.create(csap)

    form_element = form.element('form')
    form['_id'] = 'MainForm'

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = "MainForm"

    submit = form.element('input[type=submit]')

    if full_width:
        # make the inputs in the table full width
        table = form.element('table')
        table['_class'] = 'full-width'

    return dict(form=form, submit=submit)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('update', 'customers_suscriptions_alt_prices'))
def subscription_alt_price_repeat():
    csapID = request.vars['csapID']
    row = db.customers_subscriptions_alt_prices(csapID)

    cs = db.customers_subscriptions(row.customers_subscriptions_id)

    year = row.SubscriptionYear
    month = row.SubscriptionMonth

    if month == 12:
        month = 1
        year += 1
    else:
        month += 1
    db.customers_subscriptions_alt_prices.insert(customers_subscriptions_id=row.customers_subscriptions_id,
                                                 SubscriptionYear=year,
                                                 SubscriptionMonth=month,
                                                 Amount=row.Amount,
                                                 Description=row.Description,
                                                 Note=row.Note)

    redirect(URL("customers", 'subscription_alt_prices', vars={'csID':row.customers_subscriptions_id,
                                                               'cuID':cs.auth_customer_id}))


@auth.requires_login()
def subscription_alt_prices_edit():
    '''
        Edit page for subscription alternative price
    '''
    csID = request.vars['csID']
    cuID  = request.vars['cuID']
    csapID = request.vars['csapID']
    response.view = 'general/tabs_menu.html'

    customer = Customer(cuID)
    cs = CustomerSubscription(csID)
    response.title = customer.get_name()
    response.subtitle = subscription_edit_get_subtitle(csID)

    csap = db.customers_subscriptions_alt_prices

    return_url = URL('subscription_alt_prices', vars={'cuID':cuID,
                                                      'csID':csID})

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.update_next = return_url
    crud.settings.update_deletable = False
    form = crud.update(csap, csapID)

    content_back = os_gui.get_button('back_bs', return_url)
    content = DIV(content_back, form)

    menu = subscription_edit_get_menu(cuID, csID, 'subscription_alt_prices')

    back = subscription_edit_get_back(cuID)

    return dict(content=content,
                menu=menu,
                back=back,
                left_sidebar_enabled=True)


def subscription_edit_get_subtitle(csID):
    '''
        Returns subtitle for subscription edit pages
    '''
    cs = CustomerSubscription(csID)

    return SPAN(T("Edit subscription"), ' ', cs.name)


def subscription_edit_get_back(cuID):
    '''
        Returns back button for customer subscription edit pages
    '''
    return os_gui.get_button('back',
        URL('subscriptions', vars={'cuID':cuID}),
        _class='')


def subscription_edit_get_menu(cuID, csID, page):
    '''
        Returns submenu for subscription edit pages
    '''
    vars = { 'cuID':cuID,
             'csID':csID }

    pages = []

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('update', 'customers_subscriptions'):
        pages.append(['subscription_edit',
                      SPAN(os_gui.get_fa_icon('fa-edit'), ' ', T("Edit")),
                      URL('subscription_edit', vars=vars)])

        pages.append(['subscription_pauses',
                      SPAN(os_gui.get_fa_icon('fa-pause'), ' ', T("Pauses")),
                      URL('subscription_pauses', vars=vars)])

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'invoices'):
        pages.append(['subscription_invoices',
                      SPAN(os_gui.get_fa_icon('fa-file-o'), ' ', T("Invoices")),
                      URL('subscription_invoices', vars=vars)])

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('update', 'customers_subscriptions_alt_prices'):
        pages.append(['subscription_alt_prices',
                      SPAN(os_gui.get_fa_icon('fa-random'), ' ', T("Alt. Prices")),
                      URL('subscription_alt_prices', vars=vars)])

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('read', 'customers_subscriptions_credits'):
        pages.append(['subscription_credits',
                      SPAN(os_gui.get_fa_icon('fa-check-square-o'), ' ', T("Credits")),
                      URL('subscription_credits', vars=vars)])


    return os_gui.get_submenu(pages, page, horizontal=True, htype='tabs')


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_subscriptions'))
def subscriptions():
    '''
        Lists subscriptions for a customer
        request.vars['cuID'] is expected to be the customersID
    '''
    customers_id = request.vars['cuID']
    response.view = 'customers/edit_general.html'

    row = db.auth_user(customers_id)
    response.title = row.display_name
    response.subtitle = T("Subscriptions")

    header = THEAD(TR(TH('#'),
                      TH(db.customers_subscriptions.school_subscriptions_id.label),
                      TH(db.customers_subscriptions.Startdate.label),
                      TH(db.customers_subscriptions.Enddate.label),
                      TH(db.customers_subscriptions.payment_methods_id.label),
                      TH(db.customers_subscriptions.Note.label),
                      TH(T('Pauses')),
                      TH(T('Credits')),
                      TH()) # buttons
                   )

    table = TABLE(header, _class='table table-hover table-striped')

    query = (db.customers_subscriptions.auth_customer_id == customers_id)
    rows = db(query).select(db.customers_subscriptions.id,
                            db.customers_subscriptions.auth_customer_id,
                            db.customers_subscriptions.school_subscriptions_id,
                            db.customers_subscriptions.Startdate,
                            db.customers_subscriptions.Enddate,
                            db.customers_subscriptions.payment_methods_id,
                            db.customers_subscriptions.Note,
                            orderby=~db.customers_subscriptions.Startdate)

    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        delete_permission = auth.has_membership(group_id='Admins') or \
                            auth.has_permission('delete', 'customers_subscriptions')

        delete = ''
        if delete_permission:
            confirm_msg = T("Really delete this subscription?")
            onclick_del = "return confirm('" + confirm_msg + "');"
            delete = os_gui.get_button('delete_notext',
                                        URL('subscription_delete', vars={'cuID': customers_id,
                                                                         'csID': row.id}),
                                        onclick=onclick_del,
                                       _class='pull-right')

        edit = subscriptions_get_link_edit(row)


        tr = TR(TD(row.id),
                TD(repr_row.school_subscriptions_id),
                TD(repr_row.Startdate),
                TD(repr_row.Enddate),
                TD(repr_row.payment_methods_id),
                TD(repr_row.Note),
                TD(subscriptions_get_link_latest_pauses(row)),
                TD(subscriptions_get_link_credits(row)),
                TD(delete, edit))

        table.append(tr)

    add = ''
    if ( auth.has_membership(group_id='Admins') or
         auth.has_permission('create', 'customers_subscriptions') ):
        add_url = URL('subscription_add', vars={'cuID':customers_id})
        add = os_gui.get_button('add', add_url, T("Add a new subscription"), btn_size='btn-sm', _class='pull-right')

    content = DIV(table)

    back = edit_get_back()
    menu = customers_get_menu(customers_id, request.function)

    return dict(content=content, menu=menu, back=back, tools=add)


def subscriptions_get_link_edit(row):
    '''
        Returns drop down link for subscriptions
    '''
    vars = {'cuID': row.auth_customer_id,
            'csID': row.id}

    links = []

    permission = ( auth.has_membership(group_id='Admins') or
                   auth.has_permission('update', 'customers_subscriptions') )
    if permission:
        link_edit = A((os_gui.get_fa_icon('fa-pencil'), T('Edit')),
                      _href=URL('subscription_edit', vars=vars))
        links.append(link_edit)

    permission = ( auth.has_membership(group_id='Admins') or
                   auth.has_permission('update', 'customers_subscriptions_paused') )

    if permission:
        link_pauses = A((os_gui.get_fa_icon('fa-pause'), T('Pauses')),
                        _href=URL('subscription_pauses', vars=vars))
        links.append(link_pauses)

    permission = ( auth.has_membership(group_id='Admins') or
                   auth.has_permission('read', 'invoices') )
    if permission:
        link_invoices = A((os_gui.get_fa_icon('fa-file-o'), ' ', T('Invoices')),
                          _href=URL('subscription_invoices', vars=vars))

        links.append(link_invoices)

    permission = ( auth.has_membership(group_id='Admins') or
                   auth.has_permission('update', 'customers_subscriptions_alt_prices') )

    if permission:
        link = A((os_gui.get_fa_icon('fa-random'), ' ', T('Alt. prices')),
                        _href=URL('subscription_alt_prices', vars=vars))
        links.append(link)


    permission = ( auth.has_membership(group_id='Admins') or
                   auth.has_permission('read', 'customers_subscriptions_credits') )
    if permission:
        link_credits = A((os_gui.get_fa_icon('fa-check-square-o'), ' ', T("Credits")),
                         _href=URL('subscription_credits', vars=vars))
        links.append(link_credits)

    menu = os_gui.get_dropdown_menu(
        links=links,
        btn_text='',
        btn_size='btn-sm',
        btn_icon='pencil',
        menu_class='btn-group pull-right')

    return menu


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_subscriptions_credits'))
def subscription_credits_month_expired():
    '''
        Page to list given credits for a selected month for all customers
    '''
    # process request.vars
    subscription_credits_month_set_date()

    # title and subtitle
    response.title = T('Customers')
    response.subtitle = T('Subscription credits')
    response.view = 'customers/subscription_credits_month.html'

    # get from
    result = subscription_credits_month_get_form(
        session.customers_subscription_credits_month,
        session.customers_subscription_credits_year,
        current_url = request.function
    )
    response.subtitle += ' - ' + result['subtitle']

    form = result['form']
    month_chooser = os_gui.get_month_chooser(
        request.function,
        'subscription_credits_set_month',
        session.customers_subscription_credits_year,
        session.customers_subscription_credits_month
    )

    content = subscription_credits_month_get_content(expired=True)

    expire = ''
    permission = (auth.has_membership(group_id='Admins') or
                  auth.has_permission('update', 'customers_subscriptions_credits'))
    if permission:
        expire = os_gui.get_button('noicon',
                                   URL('subscription_credits_month_expire_credits'),
                                   title=T('Expire'),
                                   btn_class='btn-primary')

    back = os_gui.get_button('back', URL('customers', 'index'))
    menu = subscription_credits_month_get_menu(request.function)

    return dict(content=content,
                form=form,
                current=result['current'],
                header_tools=expire,
                month_chooser=month_chooser,
                menu=menu,
                back=back)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'customers_subscriptions_credits'))
def subscription_credits_month_expire_credits():
    '''
        Expire credits on the current day
    '''
    csch = CustomersSubscriptionsCreditsHelper()
    sub_credits_expired = csch.expire_credits(TODAY_LOCAL)

    session.flash = T('Expired credits for') + ' ' + unicode(sub_credits_expired) + ' ' + T('subscriptions')

    redirect(URL('subscription_credits_month_expired'))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_subscriptions_credits'))
def subscription_credits_month():
    '''
        Page to list given credits for a selected month for all customers
    '''
    # process request.vars
    subscription_credits_month_set_date()

    # title and subtitle
    response.title = T('Customers')
    response.subtitle = T('Subscription credits')

    # get from
    result = subscription_credits_month_get_form(
        session.customers_subscription_credits_month,
        session.customers_subscription_credits_year,
        request.function
    )
    response.subtitle += ' - ' + result['subtitle']

    form = result['form']
    month_chooser = os_gui.get_month_chooser(
        request.function,
        'subscription_credits_set_month',
        session.customers_subscription_credits_year,
        session.customers_subscription_credits_month
    )

    add = os_gui.get_button('add',
                            URL('customers', 'subscription_credits_month_add_confirm'),
                            btn_class='btn-primary')
    content = subscription_credits_month_get_content()
    back = os_gui.get_button('back', URL('customers', 'index'))
    menu = subscription_credits_month_get_menu(request.function)

    return dict(content=content,
                form=form,
                current=result['current'],
                header_tools=add,
                month_chooser=month_chooser,
                menu=menu,
                back=back)


def subscription_credits_month_get_content(expired=False):
    '''
        :param expired: Boolean
        Get list of credits of this month, default is added credits, but expired credits are returned when
        expired boolean is True
    '''
    first_day = datetime.date(session.customers_subscription_credits_year,
                              session.customers_subscription_credits_month,
                              1)
    last_day = get_last_day_month(first_day)

    left = [
        db.customers_subscriptions.on(db.customers_subscriptions_credits.customers_subscriptions_id ==
                                      db.customers_subscriptions.id),
        db.auth_user.on(db.customers_subscriptions.auth_customer_id == db.auth_user.id),
        db.school_subscriptions.on(db.customers_subscriptions.school_subscriptions_id ==
                                   db.school_subscriptions.id)
    ]

    if expired:
        query = (db.customers_subscriptions_credits.MutationType == 'sub') & \
                (db.customers_subscriptions_credits.Expiration == True) & \
                (db.customers_subscriptions_credits.MutationDateTime <= last_day) & \
                (db.customers_subscriptions_credits.MutationDateTime >= first_day)
    else:
        query = (db.customers_subscriptions_credits.MutationType == 'add') & \
                (db.customers_subscriptions_credits.MutationDateTime <= last_day) & \
                (db.customers_subscriptions_credits.MutationDateTime >= first_day)

    rows = db(query).select(db.auth_user.id,
                            db.auth_user.display_name,
                            db.auth_user.thumbsmall,
                            db.auth_user.birthday,
                            db.auth_user.archived,
                            db.school_subscriptions.Name,
                            db.customers_subscriptions.id,
                            db.customers_subscriptions.Startdate,
                            db.customers_subscriptions.school_subscriptions_id,
                            db.customers_subscriptions_credits.MutationDateTime,
                            db.customers_subscriptions_credits.MutationType,
                            db.customers_subscriptions_credits.MutationAmount,
                            left=left,
                            orderby=db.auth_user.display_name
                            )

    header = THEAD(TR(
        TH(),
        TH(T('Customer')),
        TH(T('Subscription')),
        TH(T('Subscription start')),
        TH(T('Mutation Date')),
        TH(T('Credits')),
        TH(),
        TH(),
    ))

    table = TABLE(header, _class='table table-striped')
    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        tr = TR(
            TD(repr_row.auth_user.thumbsmall, _class='os-customer_image_td'),
            TD(repr_row.auth_user.display_name),
            TD(repr_row.customers_subscriptions.school_subscriptions_id),
            TD(repr_row.customers_subscriptions.Startdate),
            TD(repr_row.customers_subscriptions_credits.MutationDateTime),
            TD(repr_row.customers_subscriptions_credits.MutationAmount),
            TD(repr_row.customers_subscriptions_credits.MutationType),
            TD(os_gui.get_button('edit',
                                 URL('subscription_credits', vars={'cuID':row.auth_user.id,
                                                                   'csID':row.customers_subscriptions.id}),
                                 _class="pull-right"))
        )

        table.append(tr)

    return table


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_subscriptions_credits'))
def subscription_credits_set_month():
    '''
        Sets the session variables for customers_subscriptions_credits year and month
    '''
    year  = request.vars['year']
    month = request.vars['month']
    back  = request.vars['back']

    session.customers_subscription_credits_year  = int(year)
    session.customers_subscription_credits_month = int(month)

    redirect(URL('subscription_credits_month'))


def subscription_credits_month_set_date(var=None):
    '''
        Set session variable for date vars
    '''
    today = TODAY_LOCAL
    if 'year' in request.vars:
        year = int(request.vars['year'])
    elif not session.customers_subscription_credits_year is None:
        year = session.customers_subscription_credits_year
    else:
        year = today.year
    session.customers_subscription_credits_year = year
    if 'month' in request.vars:
        month = int(request.vars['month'])
    elif not session.customers_subscription_credits_month is None:
        month = session.customers_subscription_credits_month
    else:
        month = today.month
    session.customers_subscription_credits_month = month


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_subscriptions_credits'))
def subscription_credits_month_show_current():
    '''
        Show current
    '''
    session.customers_subscription_credits_year  = int(TODAY_LOCAL.year)
    session.customers_subscription_credits_month = int(TODAY_LOCAL.month)

    redirect(URL('subscription_credits_month'))


def subscription_credits_month_get_form(month, year, current_url, _class='col-md-4'):
    '''
        Get month chooser form for subscription_credits_month
    '''
    months = get_months_list()

    for m in months:
        if m[0] == month:
            month_title = m[1]
    subtitle = month_title + " " + unicode(year)

    form = SQLFORM.factory(
        Field('month',
              requires=IS_IN_SET(months, zero=None),
              default=month,
              label=T("")),
        Field('year', 'integer',
              default=year,
              label=T("")),
        submit_button=T("Run report")
    )
    form.attributes['_name'] = 'form_select_date'
    form.attributes['_class'] = 'overview_form_select_date'

    input_month = form.element('select[name=month]')
    input_month.attributes['_onchange'] = "this.form.submit();"

    input_year = form.element('input[name=year]')
    input_year.attributes['_onchange'] = "this.form.submit();"
    input_year.attributes['_type'] = 'number'

    form.element('input[name=year]')

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    ## Show current
    show_current = A(T("Current month"),
                     _href=URL(current_url),
                     _class='btn btn-default')

    form = DIV(XML('<form id="MainForm" action="#" enctype="multipart/form-data" method="post">'),
               DIV(form.custom.widget.month,
                   form.custom.widget.year,
                   _class=_class),
               form.custom.end,
               _class='row')

    return dict(form=form, subtitle=subtitle, current=show_current, submit=submit)


def subscription_credits_month_get_menu(page=None):
    pages = [
        (['subscription_credits_month', T('Added credits'), URL('customers', 'subscription_credits_month')]),
        (['subscription_credits_month_expired',
          T('Expired credits'),
          URL('customers', 'subscription_credits_month_expired')]),
    ]

    return os_gui.get_submenu(pages,
                              page,
                              horizontal=True,
                              htype='tabs')


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'customers_subscriptions_credits'))
def subscription_credits_month_add_confirm():
    '''
        Show confirmation page before adding credits for a month
    '''
    response.title = T('Customers')
    response.subtitle = T('Subscription credits - Add confirmation')
    response.view = 'general/only_content_header_footer.html'

    # Set default values for session vars if not set
    subscription_credits_month_set_date()

    year = session.customers_subscription_credits_year
    month = session.customers_subscription_credits_month
    date = datetime.date(year, month, 1)

    box_title = SPAN(T('Add subscription credits for'), ' ', date.strftime('%B %Y'))

    content = DIV(

        P(T("This operation will only add credits for subscriptions where credits haven't been granted yet for this month.")),
        P(T("Paused subscriptions and subscriptions where credits are already granted for this month will be skipped."))
    )

    buttons = SPAN(
        os_gui.get_button('noicon',
                          URL('customers', 'subscription_credits_month_add'),
                          title=T('Continue'),
                          btn_class='btn-primary'),
        os_gui.get_button('noicon',
                          URL('customers', 'subscription_credits_month'),
                          title=T('Cancel'),
                          btn_class='btn-link')
    )


    back = os_gui.get_button('back', URL('customers', 'subscription_credits_month'))

    return dict(content=content,
                box_title=box_title,
                footer=buttons,
                back=back)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'customers_subscriptions_credits'))
def subscription_credits_month_add():
    '''
        Add credits for subscriptions in selected month
    '''
    from openstudio import CustomersSubscriptionsCreditsHelper

    year = session.customers_subscription_credits_year
    month = session.customers_subscription_credits_month

    csch = CustomersSubscriptionsCreditsHelper()
    added = csch.add_credits(year, month)

    session.flash = T("Added subscription credits for") + ' ' + unicode(added) + ' ' + T('customers') + '.'

    redirect(URL('subscription_credits_month'))


def payments_delete_payment_info(form):
    page = redirect(URL('payments', vars={'cuID':customers_id}))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_payments'))
def payments():
    '''
        Lists all payments a customer has made for invoices
    '''
    customers_id = request.vars['cuID']
    customer = Customer(customers_id)
    response.title = customer.get_name()
    response.subtitle = T("Payment info")
    function = request.function

    # back button
    back = edit_get_back()

    # payment_info
    db.customers_payment_info.id.readable=False
    db.customers_payment_info.auth_customer_id.readable=False
    query = (db.customers_payment_info.auth_customer_id == customers_id)
    db.customers_payment_info.auth_customer_id.default = customers_id
    fields = [ db.customers_payment_info.payment_methods_id,
               db.customers_payment_info.AccountNumber,
               db.customers_payment_info.AccountHolder,
               db.customers_payment_info.BIC,
               db.customers_payment_info.BankName,
               db.customers_payment_info.BankLocation ]
    links = [lambda row: os_gui.get_button('edit',
                                    URL('payment_info_edit',
                                        args=[customers_id, row.id])) ]

    delete_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('delete', 'customers_payment_info')

    bd_grid = SQLFORM.grid(query,
                           fields=fields,
                           links=links,
                           create=False,
                           details=False,
                           editable=False,
                           csv=False,
                           searchable=False,
                           deletable=delete_permission,
                           field_id=db.customers_payment_info.id,
                           ui=grid_ui)
    if db(query).count() == 0:
        add = os_gui.get_button('add', URL('payment_info_add', args=[customers_id]))
    else:
        add = ''
    bd_grid.element('.web2py_counter', replace=None) # remove the counter
    bd_grid.elements('span[title=Delete]', replace=None) # remove text from delete button
    bd = DIV(add, bd_grid)

    # alternative payments
    db.alternativepayments.id.readable=False
    db.alternativepayments.auth_customer_id.readable=False
    query = (db.alternativepayments.auth_customer_id == customers_id)
    db.alternativepayments.auth_customer_id.default = customers_id

    fields = [ db.alternativepayments.PaymentYear,
               db.alternativepayments.PaymentMonth,
               db.alternativepayments.Amount,
               db.alternativepayments.payment_categories_id,
               db.alternativepayments.Description ]

    links = [lambda row: os_gui.get_button('edit',
                                    URL('alternativepayment_edit',
                                        args=[customers_id, row.id])),
            lambda row: A(SPAN(_class="buttontext button", _title=T("Repeat")),
                          SPAN(_class="glyphicon glyphicon-repeat"),
                          " ", T("Repeat"),
                          _class="btn btn-default",
                          _href=URL('alternativepayment_repeat',
                                    vars=dict(apID=row.id)),
                          _title=T("Repeat")) ]
    maxtextlengths = {'alternativepayments.Description' : 30}

    delete_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('delete', 'alternativepayments')

    ap_grid = SQLFORM.grid(query,
                           fields=fields,
                           links=links,
                           maxtextlengths=maxtextlengths,
                           create=False,
                           details=False,
                           editable=False,
                           csv=False,
                           searchable=False,
                           deletable=delete_permission,
                           ui=grid_ui,
                           orderby=~db.alternativepayments.PaymentYear|\
                                   ~db.alternativepayments.PaymentMonth,
                           field_id=db.alternativepayments.id)
    add = os_gui.get_button('add', URL('alternativepayment_add', args=[customers_id]))
    ap_grid.element('.web2py_counter', replace=None) # remove the counter
    ap_grid.elements('span[title=Delete]', replace=None) # remove text from delete button
    ap = DIV(add, ap_grid)

    menu = customers_get_menu(customers_id, request.function)

    tabs = UL(LI(A(T("Payment info"),
                   _href='#bd',
                   _role='tab'),
                 _role='presentation',
                 _class='active pay_tab'),
              LI(A(T("Direct debit extra"),
                   _href='#ap',
                   _role='tab'),
                 _role='presentation',
                 _class='pay_tab'),
              _class='nav nav-tabs',
              _role='tablist',
              _id='payment_tabs')

    tab_content = DIV(DIV(bd,
                          _id='bd',
                          _role='tabpanel',
                          _class='tab-pane fade in active'),
                      DIV(ap,
                          _id='ap',
                          _role='tabpanel',
                          _class='tab-pane fade'),
                      _class='tab-content')
    content = DIV(tabs, BR(), tab_content, _role='tabpanel')

    # assign default value
    if not session.customers_payments_tab:
        session.customers_payments_tab = '#bd'

    return dict(content=content,
                menu=menu,
                back=back,
                left_sidebar_enabled=True)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_notes_backoffice') or \
               auth.has_permission('read', 'customers_notes_teachers'))
def notes():
    '''
        Lists all notes for the backoffice
        request.vars['note_type'] can be 2 values
            'backoffice' for a backoffice note
            'teacher' for a teacher note
    '''
    customers_id = request.vars['cuID']
    note_type = request.vars['note_type']
    latest = request.vars['latest']
    latest_length = request.vars['latest_length']
    try:
        latest_length = int(latest_length)
    except:
        latest_length = 50 # set default

    backoffice_class = ''
    teachers_class = ''
    all_class = ''

    active_class = 'web2py-menu-active'

    db.auth_user._format = '%(display_name)s'

    query = (db.customers_notes.auth_customer_id == customers_id)

    if note_type is None:
        db.customers_notes.BackofficeNote.default = True

    if note_type == 'backoffice':
        db.customers_notes.BackofficeNote.default = True
        query &= (db.customers_notes.BackofficeNote == True)

    if note_type == 'teachers':
        db.customers_notes.TeacherNote.default = True
        query &= (db.customers_notes.TeacherNote == True)

    notes = UL(_id='os-customers_notes')
    rows = db(query).select(db.customers_notes.ALL,
                            orderby=~db.customers_notes.NoteDate|\
                                    ~db.customers_notes.NoteTime)
    for row in rows.render():
        row_note_type = ''
        if row.BackofficeNote:
            row_note_type = T('Back office')
        elif row.TeacherNote:
            row_note_type = T('Teachers')

        if row.Alert is True:
            alert = SPAN(XML(' ! ! '),
                         _title=T("Alert"),
                         _class='red bold')
        else:
            alert = ''

        if latest == 'True':
            note = DIV(alert,
                       XML(max_string_length(row.Note.replace('\n','<br>'),
                                             latest_length)))
            break

        else:
            buttons = DIV(_class='btn-group pull-right')
            if auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'customers_notes'):
                edit = os_gui.get_button('edit_notext',
                                  URL('note_edit', args=[row.id]),
                                  cid=request.cid)
                buttons.append(edit)

            if auth.has_membership(group_id='Admins') or \
               auth.has_permission('delete', 'customers_notes'):
                remove = os_gui.get_button('delete_notext', '#')
                buttons.append(remove)

            # correct time for timezone
            #TODO: Move notedate and notetime fields into notedatetime and then represent using pytz


            notes.append(LI(buttons,
                            alert,
                            SPAN(row.NoteDate,
                                 ' ',
                                 row.NoteTime,
                                 _class='bold'),
                            SPAN(' - ',
                                 row.auth_user_id,
                                 _class='grey'),
                            BR(),
                            XML(row.Note.replace('\n','<br>')),
                            _id='note_' + unicode(row.id)))

    if latest == 'True':
        try:
            return_value = note
        except:
            # no rows found
            return_value = ''
    else:
        vars = {'cuID':customers_id}
        if not note_type or note_type == 'backoffice':
            vars['note_type'] = 'backoffice'
        else:
            vars['note_type'] = 'teachers'

        perm = auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'customers_notes')
        if perm:
            add = notes_get_add()
            add_title = H4(T('Add a new note'))
        else:
            add = ''
            add_title = ''

        content = DIV(add_title,
                      add,
                      notes)

        return_value = dict(content=content)

    response.js = "iconHandlers()"

    return return_value


@auth.requires_login()
def note_edit():
    '''
        Provides an edit page for a note.
        request.args[0] is expected to be the customers_note_id (cn_id)
    '''
    cn_id = request.args[0]

    note = db.customers_notes(cn_id)
    customers_id = note.auth_customer_id

    if note.BackofficeNote:
        note_type = 'backoffice'
    elif note.TeacherNote:
        note_type = 'teachers'

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T('Saved')
    form = crud.update(db.customers_notes, cn_id)

    form.custom.widget.Note['_class'] += ' form-control'

    form = DIV(form.custom.begin,
               form.custom.widget.Note,
               form.custom.submit,
               form.custom.end,
               _class='os-customers_notes_edit clear')

    back =  os_gui.get_button('back',
                      URL('notes', vars={'cuID':customers_id,
                                         'note_type':note_type}),
                      _class='left',
                      cid=request.cid)

    content = DIV(BR(),
                  back,
                  BR(),BR(),
                  form)

    response.js = "setTimeout(function() { $('div.flash').fadeOut(); }, 2500 );"

    return dict(content=content)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('delete', 'customers_notes'))
def note_delete():
    '''
        Called as JSON, used to remove a note
    '''
    response.view = 'generic.json'
    cn_id = request.vars['id']

    status = 'success'
    message = T('Deleted note')

    query = (db.customers_notes.id == cn_id)
    if not db(query).delete(): # returns 0 when it fails
        status = 'fail'
        message = T("Delete failed")

    return dict(status=status, message=message)


def notes_get_add(var=None):
    '''
        Provides a page to add a note
        request.vars['note_type'] can be 2 values
            'backoffice' for a backoffice note
            'teachers' for a teachers note
    '''
    note_type = request.vars['note_type']
    customers_id = request.vars['cuID']

    if note_type is None:
        vars = {'cuID':customers_id}
    else:
        vars = {'cuID':customers_id,
                'note_type':note_type}

    return_url = URL('notes', vars=vars)

    db.customers_notes.auth_customer_id.default = customers_id
    db.customers_notes.auth_user_id.default = auth.user.id


    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T('')
    crud.settings.create_next = return_url
    form = crud.create(db.customers_notes)

    form.custom.widget.Note['_class'] += ' form-control'

    form = DIV(form.custom.begin,
               form.custom.widget.Note,
               form.custom.submit,
               form.custom.end,
               _class='os-customers_notes_edit')

    return form


def payment_info_get_returl_url(customers_id):
    '''
        Returns the return url for payment_info_add and payment_info_edit
    '''
    return URL('payments', vars={'cuID':customers_id})


@auth.requires_login()
def payment_info_add():
    """
        This function shows an add page for payment_info
        request.args[0] is expected to be the customers_id
    """
    customers_id = request.args[0]
    customer = Customer(customers_id)
    response.title = customer.get_name()
    response.subtitle = T("Payment info")
    response.view = 'general/only_content.html'

    session.customers_payments_tab = '#bd'

    return_url = payment_info_get_returl_url(customers_id)

    submit = ''
    query = (db.customers_payment_info.auth_customer_id == customers_id)
    if db(query).count() < 1:
        db.customers_payment_info.auth_customer_id.default = customers_id

        crud.messages.submit_button = T("Save")
        crud.messages.record_created = T("Added payment info")
        crud.settings.create_next = return_url
        form = crud.create(db.customers_payment_info)

        result = set_form_id_and_get_submit_button(form, 'MainForm')
        form = result['form']
        submit = result['submit']
    else:
        form = T("Payment info already entered for this customer.")

    back = os_gui.get_button('back', return_url)

    return dict(content=form, save=submit, back=back)


@auth.requires_login()
def payment_info_edit():
    """
        This function shows an edit page for the payment_info of a customer
        request.args[0] is expected to be the customers_id
        request.args[1] is expected to be the payment_infoID (piID)
    """
    customers_id = request.args[0]
    piID = request.args[1]
    customer = Customer(customers_id)
    response.title = customer.get_name()
    response.subtitle = T("Payment info")
    response.view = 'general/only_content.html'

    session.customers_payments_tab = '#bd'

    return_url = payment_info_get_returl_url(customers_id)

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Updated payment info")
    crud.settings.update_next = return_url
    crud.settings.update_deletable = False
    form = crud.update(db.customers_payment_info, piID)

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    back = os_gui.get_button('back', return_url)

    return dict(content=form, save=submit, back=back)


def alternativepayment_get_return_url(customers_id):
    '''
        Returns return url for alternative payments
    '''
    return URL('payments', vars={'cuID':customers_id})


@auth.requires_login()
def alternativepayment_add():
    """
        This function shows an add page for alternative payments
        request.args[0] is expected to be the customers_id
    """
    customers_id = request.args[0]
    response.title = T("Payment info")
    customer = Customer(customers_id)
    response.subtitle = SPAN(T('Direct debit extra'), ' ', customer.get_name())
    response.view = 'general/only_content.html'

    session.customers_payments_tab = '#ap'

    db.alternativepayments.auth_customer_id.default = customers_id

    return_url = alternativepayment_get_return_url(customers_id)

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Saved")
    crud.settings.create_next = return_url
    form = crud.create(db.alternativepayments)

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    back = os_gui.get_button('back', return_url)

    return dict(content=form, save=submit, back=back)


@auth.requires_login()
def alternativepayment_edit():
    """
        This function shows an add page for alternative payments
        request.args[0] is expected to be the customers_id
        request.args[1] is expected to be the alternativepaymentsID
    """
    response.title = T("Payment info")
    customers_id = request.args[0]
    apID = request.args[1]
    customer = Customer(customers_id)
    response.subtitle = SPAN(T("Direct debit extra"), ' ', customer.get_name())
    response.view = 'general/only_content.html'

    session.customers_payments_tab = '#ap'

    db.alternativepayments.auth_customer_id.default = customers_id

    return_url = alternativepayment_get_return_url(customers_id)

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.update_next = return_url
    crud.settings.update_deletable = False
    form = crud.update(db.alternativepayments, apID)

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    back = os_gui.get_button('back', return_url)

    return dict(content=form, save=submit, back=back)


def alternativepayments_check(form):
    month = form.vars['PaymentMonth']
    year = form.vars['PaymentYear']
    categoryID = form.vars['PaymentCategoryID']

    query = (db.alternativepayments.auth_customer_id==customers_id) & \
            (db.alternativepayments.PaymentYear==year) & \
            (db.alternativepayments.PaymentMonth==month) & \
            (db.alternativepayments.payment_categories_id==categoryID)
    count = db(query).count()
    if count > 1:
        session.flash = T("Only one alternative payment per category per month is allowed.")
        db.rollback() # undo changes, we don't want this data in the db to prevent export issues.


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'workshops'))
def events():
    """
        This function shows a page which lists the workshop attendance for a
        customer.
    """
    cuID = request.vars['cuID']

    session.customers_payment_back = None
    # To redirect back after removing a product
    session.events_ticket_sell_back = 'customers'
    # Invoice back redirects
    session.invoices_edit_back = 'customers_events'
    session.invoices_payment_add_back = 'customers_events'
    # To redirect back here after sending info mail
    session.workshops_product_resend_info_mail = 'customers_events'


    session.workshops_payment_back = 'customer'
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Events")

    rows = customer.get_workshops_rows()

    wh = WorkshopsHelper()
    header = THEAD(TR(TH(_class='workshop_image_th'),
                      TH(T('Event'),
                      TH(T('Invoice')),
                      TH(T('Event info')),
                      TH(),

                      )))
    table = TABLE(header, _class="table table-hover table-striped")
    ih = InvoicesHelper()

    for i, row in enumerate(rows):
        repr_row = list(rows[i:i+1].render())[0]

        wsID     = row.workshops.id
        wsp_cuID = row.workshops_products_customers.id

        # invoice
        if row.invoices.id:
            invoice = ih.represent_invoice_for_list(
                row.invoices.id,
                repr_row.invoices.InvoiceID,
                repr_row.invoices.Status,
                row.invoices.Status,
                row.invoices.payment_methods_id
            )
        else:
            invoice = ''


        # Event info link
        link_text = T('Send')
        if row.workshops_products_customers.WorkshopInfo:
            link_text = T('Resend')
        resend_link = A(link_text, ' ', T('info mail'),
                        _href=URL('events', 'ticket_resend_info_mail', vars={'wspcID':wsp_cuID}))
        event_info = wh.get_customer_info(wsp_cuID,
                                          wsID,
                                          row.workshops_products_customers.WorkshopInfo,
                                          resend_link)

        permission = auth.has_membership(group_id='Admins') or \
                     auth.has_permission('delete', 'workshops_products_customers')
        if permission:
            confirm_remove_msg = T("Really remove this workshop product?")
            btn_delete = os_gui.get_button(
                'delete_notext',
                URL('events', 'ticket_delete_customer',
                    vars={'wsID'     : wsID,
                          'wsp_cuID' : wsp_cuID}),
                tooltip=T('Remove customer from list'),
                onclick="return confirm('" + confirm_remove_msg + "');" )
        else:
            btn_delete = ''

        # check waitinglist
        if row.workshops_products_customers.Waitinglist:
            waitinglist = os_gui.get_label('danger', T('Waitinglist'))
        else:
            waitinglist = ''

        # check full workshop
        if row.workshops_products.FullWorkshop:
            label_class = 'primary'
        else:
            label_class = 'default'

        # check cancelled
        through_class= ''
        cancelled_label = ''
        title_cancel = T('Cancel customer')
        if row.workshops_products_customers.Cancelled:
            title_cancel = T('Undo cancellation')
            cancelled_label = SPAN(' ', os_gui.get_label('warning',
                                                         T("Cancelled")))
            through_class = 'line-through'

        btn_cancel = os_gui.get_button(
            'cancel_notext',
            URL('events', 'ticket_cancel_customer',
                               vars={'wsID'     : wsID,
                                     'wsp_cuID' : wsp_cuID}),
            tooltip=title_cancel )


        tr = TR(TD(repr_row.workshops.thumbsmall),
                TD(SPAN(A(row.workshops.Name,
                          _href=URL('events', 'tickets',
                                    vars={'wsID':wsID})),
                        _class=' ' + through_class),
                   SPAN(' - ', row.workshops.Startdate,
                        _class='small_font grey'),
                   BR(),
                   os_gui.get_label(label_class, row.workshops_products.Name),
                   ' ', waitinglist, cancelled_label),
                TD(invoice),
                TD(event_info),
                TD(DIV(btn_cancel, btn_delete, _class='btn-group pull-right')))

        table.append(tr)

    # back button
    back = edit_get_back()
    add  = os_gui.get_button('add',
                             URL('event_add', vars={'cuID':cuID}),
                             btn_size='btn-sm',
                             tooltip=T("Add a workshop"))
    menu = customers_get_menu(cuID, request.function)
    loader = os_gui.get_ajax_loader(message=T("Refreshing list..."))

    return dict(content=table,
                menu=menu,
                add=add,
                back=back,
                loader=loader,
                left_sidebar_enabled=True)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('update', 'workshops_products_customers'))
def event_add():
    """
        Select a workshop to list products from
    """
    response.view = 'general/only_content.html'

    cuID = request.vars['cuID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = SPAN(T("Add event product"), B(' > '),
                             T("List events"))


    # list of workhsops
    grid = events_add_get_list(cuID)

    back = os_gui.get_button('back', URL('events',
                                         vars={'cuID':cuID}))


    return dict(content=grid,
                back=back)


def events_add_get_list(cuID):
    """
        Returns grid of workshops with select button
    """
    query = (db.workshops.Archived == False)

    links = [ lambda row: os_gui.get_button(
                        'noicon',
                        URL('events_add_list_tickets',
                            vars={'wsID' : row.id,
                            'cuID' : cuID}),
                        title=T('Products'),
                        tooltip=T('Show products for workshop')) ]

    fields = [ db.workshops.thumbsmall,
               db.workshops.Name,
               db.workshops.Startdate,
               db.workshops.Teacher,
               db.workshops.auth_teacher_id ]

    db.workshops.id.readable = False

    maxtextlengths = {'workshops.Name' : 40}
    grid = SQLFORM.grid(query,
                        fields=fields,
                        links=links,
                        details=False,
                        searchable=False,
                        csv=False,
                        create=False,
                        editable=False,
                        deletable=False,
                        maxtextlengths=maxtextlengths,
                        orderby=db.workshops.Startdate|db.workshops.Name,
                        ui = grid_ui)
    grid.element('.web2py_counter', replace=None) # remove the counter
    grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    return grid


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'workshops_products_customers'))
def events_add_list_tickets():
    '''
        List products for a workshop
    '''
    response.title = T("Add workshop product")
    response.view = 'general/only_content.html'

    # To redirect back when adding a workshop product
    session.events_ticket_sell_back = 'customers'

    cuID = request.vars['cuID']
    wsID = request.vars['wsID']
    workshop = db.workshops(wsID)
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = SPAN(T("Add event product"), B(' > '),
                             T("Products for "),
                             workshop.Name)

    # list of workhsops
    grid = events_add_list_products_get_list(wsID, cuID)

    back = os_gui.get_button('back', URL('workshops_add',
                                         vars={'cuID':cuID}))

    return dict(content=grid,
                back=back)


def events_add_list_products_get_list(wsID, cuID):
    """
        Returns list of products for a workshop
    """
    query = (db.workshops_products.workshops_id == wsID)

    links = [event_add_list_products_get_list_get_link_add]

    fields = [ db.workshops_products.id,
               db.workshops_products.Name,
               db.workshops_products.Description,
               db.workshops_products.Price ]

    left = [ db.workshops.on(db.workshops_products.workshops_id == \
                             db.workshops.id) ]

    db.workshops_products.id.readable = False

    maxtextlengths = {'workshops.Name' : 40}
    grid = SQLFORM.grid(query,
                        fields=fields,
                        left=left,
                        links=links,
                        details=False,
                        searchable=False,
                        csv=False,
                        create=False,
                        editable=False,
                        deletable=False,
                        maxtextlengths=maxtextlengths,
                        orderby=db.workshops_products.FullWorkshop | \
                                db.workshops_products.Name,
                        ui = grid_ui)
    grid.element('.web2py_counter', replace=None) # remove the counter
    grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    return grid


def event_add_list_products_get_list_get_link_add(row):
    """
        Returns an add button if a customer isn't already added,
        otherwise returns ''
    """
    cuID  = request.vars['cuID']
    wspID = row.workshops_products.id
    row  = db.workshops_products(wspID)
    wsID = row.workshops_id

    wh = WorkshopsHelper()

    buttons = wh.get_product_sell_buttons(cuID, wsID, wspID, request.cid)

    return buttons


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'customers_documents'))
def documents():
    '''
        This function shows a list of documents uploaded for a customer
    '''
    customers_id = request.vars['cuID']
    response.view = 'customers/edit_general.html'
    customer = Customer(customers_id)
    response.title = customer.get_name()
    response.subtitle = T("Documents")

    db.customers_documents.id.readable=False
    db.customers_documents.UploadDateTime.readable=True
    # set the default value in the grid
    links = [lambda row: os_gui.get_button('edit',
                                     URL('document_edit',
                                         args=[customers_id, row.id]),
                                     T("Edit document description"))]

    query = (db.customers_documents.auth_customer_id == customers_id)
    maxtextlengths = {'customers_documents.Description' : 30}
    headers = {'customers_documents.DocumentFile': T("File")}

    fields = [db.customers_documents.UploadDateTime,
              db.customers_documents.Description,
              db.customers_documents.DocumentFile]

    delete_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('delete', 'customers_documents')
    grid = SQLFORM.grid(query, fields=fields, links=links,
        headers=headers,
        details=False,
        searchable=False,
        deletable=delete_permission,
        csv=False,
        create=False,
        editable=False,
        maxtextlengths=maxtextlengths,
        orderby=~db.customers_documents.UploadDateTime,
        field_id=db.customers_documents.id,
        ui = grid_ui)
    grid.element('.web2py_counter', replace=None) # remove the counter
    grid.elements('span[title=Delete]', replace=None) # remove text from delete button

    add_url = URL('document_add', args=[customers_id])
    add = os_gui.get_button('add', add_url, T("Add a new document"),  btn_size='btn-sm', _class='pull-right')

    back = edit_get_back()

    menu = customers_get_menu(customers_id, request.function)

    return dict(content=grid, menu=menu, back=back, add=add)


def documents_get_return_url(customers_id):
    '''
        Returns the return url for documents add and edit
    '''
    return URL('documents', vars={'cuID':customers_id})


@auth.requires_login()
def document_add():
    """
        This function shows an add page for a document
        request.args[0] is expected to be the customers_id
    """
    customers_id = request.args[0]
    response.view = 'general/only_content.html'
    customer = Customer(customers_id)
    response.title = customer.get_name()
    response.subtitle = T("Upload document")

    db.customers_documents.auth_customer_id.default = customers_id

    return_url = documents_get_return_url(customers_id)

    space = uploads_available_space(request.folder)
    if space['available'] > 4:
        crud.messages.submit_button = T("Save")
        crud.messages.record_created = T("Uploaded document")
        crud.settings.create_next = return_url
        form = crud.create(db.customers_documents)

        form_element = form.element('form')
        form['_id'] = 'MainForm'

        elements = form.elements('input, select, textarea')
        for element in elements:
            element['_form'] = "MainForm"

        submit = form.element('input[type=submit]')

        content = form
    else:
        content = space['full_message']

    back = os_gui.get_button("back", return_url)

    return dict(content=content, back=back, save=submit)


@auth.requires_login()
def document_edit():
    """
        This function shows an edit page for a document
        request.args[0] is expected to be the customers_id
        request.args[1] is expected to be the customers_documentID (cudID)
    """
    response.view = 'general/only_content.html'

    customers_id = request.args[0]
    cudID = request.args[1]
    customer = Customer(customers_id)
    response.title = customer.get_name()
    response.subtitle = T("Edit Document")

    db.customers_documents.DocumentFile.readable = False
    db.customers_documents.DocumentFile.writable = False

    return_url = documents_get_return_url(customers_id)

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Updated document")
    crud.settings.update_next = return_url
    crud.settings.update_deletable = False
    form = crud.update(db.customers_documents, cudID)

    form_element = form.element('form')
    form['_id'] = 'MainForm'

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = "MainForm"

    submit = form.element('input[type=submit]')

    back = os_gui.get_button("back", return_url)

    return dict(content=form, back=back, save=submit)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'auth_user'))
def load_list_set_search():
    '''
        Expected to be called as JSON

        variables that can be set are:
        - name
    '''
    name = request.vars['name']
    session.customers_load_list_search_name = '%' + name.strip() + '%'
    try:
        search_id = int(name.strip())
        session.customers_load_list_search_name_int = search_id
    except ValueError:
        session.customers_load_list_search_name_int = None

    message = T("Updated search requirements")
    status = 'success'

    return dict(message=message, status=status)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'auth_user'))
def load_list():
    '''
        Returns a list of customers, to be used as LOAD
        request.vars['items_per_page'] sets the items shown on each page
        request.vars['list_type'] can be 'classes_attendance_list'
        request.vars['show_location'] can be 'True' or 'False'
    '''
    items_per_page = request.vars['items_per_page']
    list_type = request.vars['list_type']
    initial_list = request.vars['initial_list']
    clsID = request.vars['clsID']
    date_formatted = request.vars['date']
    wsID = request.vars['wsID']
    wspID = request.vars['wspID']

    show_location = request.vars['show_location']
    if show_location == 'True':
        show_location = True
    else:
        show_location = False

    pictures = request.vars['pictures']
    if pictures == 'True' or pictures is None:
        pictures = True
    else:
        pictures = False

    contact_permission = ( auth.has_membership(group_id='Admins') or
                           auth.has_permission('update', 'customers_contact') )

    show_email = request.vars['show_email']
    if show_email == 'True' and contact_permission:
        show_email = True
    else:
        show_email = False

    archived = request.vars['archived']
    if archived == 'False' or archived is None:
        archived = False
    else:
        archived = True

    if date_formatted:
        date = datestr_to_python(DATE_FORMAT, date_formatted)
    else:
        date = datetime.date.today()

    # general settings
    try:
        items_per_page = int(items_per_page)
    except ValueError:
        items_per_page = 10

    if len(request.args):
        page=int(request.args[0])
    else:
        page=0
    limitby=(page*items_per_page,(page+1)*items_per_page+1)

    # query and select
    search_name = session.customers_load_list_search_name

    if (search_name and search_name != '%%') or (initial_list):
        query = (db.auth_user.id > 1) & (db.auth_user.trashed == False)
    else:
        query = (db.auth_user.id < 1)

    title = ''
    if list_type == 'customers_index':
        query &= (db.auth_user.archived == archived)

    elif list_type == 'classes_attendance_list':
        title = H4(T('Search results'))
        query &= (db.auth_user.archived == False)

    elif list_type == 'classes_manage_reservation':
        query &= (db.auth_user.archived == False)

    elif list_type == 'events_ticket_sell':
        #table_class += ' full-width'
        query &= (db.auth_user.archived == False)

    if search_name:
        query &= ((db.auth_user.display_name.like(search_name)) |
                  (db.auth_user.email == search_name.replace('%', '')) |
                  (db.auth_user.id == session.customers_load_list_search_name_int))


    if initial_list and (search_name == '%%' or not search_name):
        orderby = ~db.auth_user.id
    else:
        orderby = db.auth_user.display_name


    rows = db(query).select(db.auth_user.id,
                            db.auth_user.archived,
                            db.auth_user.thumbsmall,
                            db.auth_user.birthday,
                            db.auth_user.display_name,
                            db.auth_user.email,
                            db.auth_user.school_locations_id,
                            limitby=limitby,
                            orderby=orderby)
    table_class = 'table table-hover'
    table = TABLE(_class=table_class)
    for i, row in enumerate(rows.render()):
        if i == items_per_page:
            break

        cuID = row.id

        # get subscription for customer
        customer = Customer(cuID)
        if list_type == 'selfcheckin_checkin':
            # For self check-in display of subscriptions can be configured by the user.
            show_subscriptions_prop = 'selfcheckin_show_subscriptions'
            show_subscriptions = get_sys_property(show_subscriptions_prop)
            if show_subscriptions:
                subscriptions = True
            else:
                subscriptions = False

            subscr_cards = customer.get_subscriptions_and_classcards_formatted(date,
                                                                               show_subscriptions=subscriptions)
        else:
            subscr_cards = customer.get_subscriptions_and_classcards_formatted(date)


        # get email if requested
        email = ''
        if show_email:
            email = TD(os_gui.get_fa_icon('fa-envelope-o'), ' ',
                       A(row.email,
                         _href='mailto:' + row.email,
                         _class='grey'),
                       _class='grey small_font')

        # get location if requested
        location = ''
        if show_location:
            repr_row = list(rows[i:i + 1].render())[0]
            location = TD(os_gui.get_label('info', repr_row.school_locations_id))


        # add everything to the table cell
        customer_info = TD(SPAN(row.display_name,
                                _class='customer_name bold'),
                           BR(),
                           subscr_cards)
        if pictures:
            table_row = TR(TD(row.thumbsmall, _class='os-customer_image_td'),
                           customer_info,
                           email,
                           location)

        else:
            table_row = TR(customer_info,
                           email,
                           location)


        if list_type == 'customers_index':
            buttons = TD(load_list_get_customer_index_buttons(row))
        elif list_type == 'classes_attendance_list':
            buttons = TD(load_list_get_attendance_list_buttons(row,
                                                               clsID,
                                                               date_formatted),
                         _class='table-vertical-align-middle')
        elif list_type == 'selfcheckin_checkin':
            buttons = TD(load_list_get_selfcheckin_checkin_buttons(
                row,
                clsID,
                date_formatted),
                _class='table-vertical-align-middle')
        elif list_type == 'classes_manage_reservation':
            buttons = TD(load_list_get_reservation_list_buttons(
                row,
                clsID,
                date_formatted),
                _class='table-vertical-align-middle')
        elif list_type == 'events_ticket_sell':
            buttons = TD(load_list_get_events_ticket_sell_buttons(row,
                                                                  wsID,
                                                                  wspID),
                         _class='table-vertical-align-middle')

        try:
            table_row.append(buttons)
        except:
            pass

        table.append(table_row)



    # Navigation
    previous = ''
    url_previous = None
    if page:
        url_previous = URL(args=[page-1], vars=request.vars)
        previous = A(SPAN(_class='glyphicon glyphicon-chevron-left'),
                     _href=url_previous,
                     cid=request.cid)

    nxt = ''
    url_next = None
    if len(rows) > items_per_page:
        url_next = URL(args=[page+1], vars=request.vars)
        nxt = A(SPAN(_class='glyphicon glyphicon-chevron-right'),
                _href=url_next,
                cid=request.cid)

    navigation = os_gui.get_page_navigation_simple(url_previous, url_next, page+1, request.cid)

    if previous or nxt:
        pass
    else:
        navigation = ''


    content = DIV(title, table, navigation)

    if len(rows) == 0:
        content = DIV(BR(), T("No results..."), BR(),
                      _class='grey col-md-12')

    return dict(content=content)


def load_list_get_reservation_list_buttons(row,
                                           clsID,
                                           date_formatted):
    '''
        Returns buttons for the manage_reservation list type
    '''
    cuID = row.id
    date = datestr_to_python(DATE_FORMAT, date_formatted)
    customer = Customer(cuID)
    customer_name = customer.get_name()

    recurring_url = URL('classes', 'reservation_add',
                        vars={'cuID'  : cuID,
                              'clsID' : clsID,
                              'date'  : date_formatted},
                        extension='')

    button = os_gui.get_button('add', recurring_url,
                               tooltip=T('Enroll this customer in this class'),
                               _class='pull-right')

    return button


def load_list_get_customer_index_buttons(row):
    '''
        Returns buttons for customers.py/index
    '''
    buttons = DIV(_class='btn-group')

    btn_mail = ''
    contact_permission = ( auth.has_membership(group_id='Admins') or
                           auth.has_permission('update', 'customers_contact') )

    if contact_permission:
        btn_mail = A(I(_class="fa fa-envelope"), " ",
                     _class="btn btn-default btn-sm",
                     _href='mailto:' + row.email,
                     _title=T("Mail customer"))

    buttons.append(btn_mail)

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('update', 'auth_user'):

       btn_edit =  os_gui.get_button('edit', URL('edit',
                                                  args=[row.id],
                                                  extension=''))

       buttons.append(btn_edit)


       btn_archive = index_get_link_archive(row)
       buttons.append(btn_archive)

    delete = ''
    if (auth.has_membership(group_id='Admins') or
        auth.has_permission('delete', 'auth_user')):

        if auth.user.id == row.id:
            onclick = "alert('Unable to delete, you are currently logged in using this account.');"
            url = '#'
        else:
            onclick = "return confirm('" + \
                 T('Do you really want to delete this customer?') + "');"
            url = URL('delete', vars={'cuID':row.id}, extension='')

        delete = os_gui.get_button('delete_notext',
                                   url,
                                   onclick=onclick)

    return DIV(buttons, delete, _class='pull-right')


def load_list_get_attendance_list_buttons(row,
                                          clsID,
                                          date_formatted):
    '''
        Returns buttons for the attendance_list list type
    '''
    date = datestr_to_python(DATE_FORMAT, date_formatted)
    check = db.classes_attendance(auth_customer_id=row.id,
                                  classes_id=clsID,
                                  ClassDate=date)
    if not check:
        url = URL('classes', 'attendance_booking_options',
                  vars={'cuID':row.id,
                        'clsID':clsID,
                        'date':date_formatted},
                  extension='')

        return A(T('Check in'),
                 _href=url,
                 _class='btn btn-default btn-sm pull-right')
    else:
        return ''

    # ah = AttendanceHelper()
    # cuID = row.id
    # date = datestr_to_python(DATE_FORMAT, date_formatted)
    #
    # return ah.get_signin_buttons(clsID, date, cuID)


def load_list_get_selfcheckin_checkin_buttons(row,
                                              clsID,
                                              date_formatted):
    '''
        Returns buttons for the selfcheckin_checkin list type
    '''
    date = datestr_to_python(DATE_FORMAT, date_formatted)
    check = db.classes_attendance(auth_customer_id=row.id,
                                  classes_id=clsID,
                                  ClassDate=date)
    if not check:
        url = URL('selfcheckin', 'checkin_booking_options',
                  vars={'cuID':row.id,
                        'clsID':clsID,
                        'date':date_formatted},
                  extension='')

        return A(T('Check in'),
                 _href=url,
                 _class='btn btn-default btn-sm pull-right')
    else:
        return ''


def load_list_get_events_ticket_sell_buttons(row,
                                             wsID,
                                             wspID):
    '''
        Returns buttons for workshop_product_sell list type
        This is a select button to select a customer to sell a product to
    '''
    cuID = row.id

    wh = WorkshopsHelper()

    buttons = wh.get_product_sell_buttons(cuID, wsID, wspID, request.cid)

    return buttons


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'tasks'))
def tasks():
    '''
        Display list of tasks for a customer
    '''
    response.view = 'customers/edit_general.html'
    cuID = request.vars['cuID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Tasks")
    #session.tasks_index_filter = 'open'
    cuID = request.vars['cuID']

    content = DIV(LOAD('tasks', 'list_tasks.load',
                       vars=request.vars,
                       content=os_gui.get_ajax_loader()))

    # Add permission
    add = ''
    permission = auth.has_membership(group_id='Admins') or \
                 auth.has_permission('create', 'tasks')
    if permission:
        #add = os_gui.get_button('add', url_add)
        th = TasksHelper()
        add = th.add_get_modal({'cuID':cuID})

    back = edit_get_back()
    menu = customers_get_menu(cuID, request.function)

    #save = os_gui.get_submit_button('task_edit')

    return dict(content=content,
                menu=menu,
                back=back,
                add=add)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'auth_user_account'))
def account():
    '''
        Account options for an account
    '''
    response.view = 'customers/edit_general.html'
    cuID = request.vars['cuID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Account")

    for field in db.auth_user:
        field.readable = False
        field.writable = False

    db.auth_user.archived.readable = True
    db.auth_user.archived.writable = True
    db.auth_user.customer.readable = True
    db.auth_user.customer.writable = True
    db.auth_user.enabled.readable = True
    db.auth_user.enabled.writable = True
    db.auth_user.teacher.readable = True
    db.auth_user.teacher.writable = True
    db.auth_user.employee.readable = True
    db.auth_user.employee.writable = True
    db.auth_user.business.readable = True
    db.auth_user.business.writable = True
    db.auth_user.login_start.readable = True
    db.auth_user.login_start.writable = True

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.update_deletable = False
    form = crud.update(db.auth_user, cuID)

    form_id = "MainForm"
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    submit = form.element('input[type=submit]')


    submenu = account_get_submenu(request.function, cuID)
    content = DIV(submenu, BR(), form)

    back = edit_get_back()
    menu = customers_get_menu(cuID, 'account')

    return dict(content=content,
                menu=menu,
                back=back,
                save=submit)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('merge', 'auth_user'))
def account_merge():
    '''
        Page to allow merging of 2 entries in auth_user
    '''
    response.view = 'customers/edit_general.html'
    cuID = request.vars['cuID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Merge account")

    warning = ''
    if 'auth_merge_id' in request.vars:
        auth_merge_id = request.vars['auth_merge_id']
        btn_onclick = "return confirm('" + \
                      T('Absolutely sure you want to merge the selected account into this one?') + \
                      "');"
        btn_title = XML(SPAN(T("Click here")))
        btn = A(btn_title,
                _href=URL('account_merge_execute',
                          vars={'cuID':cuID,
                                'auth_merge_id':auth_merge_id}),
                _title=T('Execute merge'),
                _class='btn btn-link',
                _id   ='btn_merge_execute',
                _onclick = btn_onclick)
        msg_extra = SPAN(T('to merge all info except the profile of account with ID'), ' ',
                         auth_merge_id, ' ',
                         T("into this account."))
        warning_text = DIV(B(T("Warning")), ' ',
                           T("Merging cannot be undone!"), ' ',
                           btn,
                           msg_extra)
        warning = DIV(HR(),
                      os_gui.get_alert('warning',
                                       warning_text,
                                       dismissable=False))


    submenu = account_get_submenu(request.function, cuID)
    description = DIV(B(T("Select an account to merge into this one")))
    form = account_merge_get_input_form(request.vars['auth_merge_id'])

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    content = DIV(submenu, BR(),
                  description,
                  form,
                  warning)

    back = edit_get_back()
    menu = customers_get_menu(cuID, 'account')

    return dict(content=content,
                menu=menu,
                back=back,
                save=submit)


def account_merge_get_input_form(auth_merge_id):
    '''
        Simple input form to enter an auth_user_id to merge with
    '''
    db.auth_user._format = '%(id)s - %(display_name)s'

    merge_query = (db.auth_user.merged == False) & \
                  (db.auth_user.id > 1) # exclude admin user
    form = SQLFORM.factory(
        Field('auth_merge_id', db.auth_user,
              default  = auth_merge_id,
              requires = IS_IN_DB(db(merge_query),
                                  'auth_user.id',
                                  '%(id)s - %(display_name)s - %(email)s - Archived: %(archived)s - Teacher: %(teacher)s',
                              zero=T("Please select...")),
              label    = T('')),
        submit_button = T('Select'),
        formstyle='table3cols'
        )

    select = form.element('#no_table_auth_merge_id')


    return form


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('merge', 'auth_user'))
def account_merge_execute():
    '''
        Actually merge account
    '''
    cuID          = request.vars['cuID'] # merge into
    auth_merge_id = request.vars['auth_merge_id'] # merge from

    merge_into = db.auth_user(cuID)
    merge_from = db.auth_user(auth_merge_id)
    if not merge_from.merged:
        # loop over all tables in db
        for table in db:
            # auth_user_id
            try:
                query = (table.auth_user_id == auth_merge_id)
                rows = db(query).select(table.ALL)
                for row in rows:
                    row.auth_user_id = cuID
                    row.update_record()
            except AttributeError:
                pass
            # auth_teacher_id
            try:
                query = (table.auth_teacher_id == auth_merge_id)
                rows = db(query).select(table.ALL)
                for row in rows:
                    row.auth_teacher_id = cuID
                    row.update_record()
            except AttributeError:
                pass
            # auth_teacher_id2
            try:
                query = (table.auth_teacher_id2 == auth_merge_id)
                rows = db(query).select(table.ALL)
                for row in rows:
                    row.auth_teacher_id2 = cuID
                    row.update_record()
            except AttributeError:
                pass
            # auth_customer_id
            try:
                query = (table.auth_customer_id == auth_merge_id)
                rows = db(query).select(table.ALL)
                for row in rows:
                    row.auth_customer_id = cuID
                    row.update_record()
            except AttributeError:
                pass

        # mark row as merged
        # set merged for auth_user auth_merge_id
        merge_from.merged   = True
        merge_from.archived = True
        # set merged_with for auth_user auth_merge_id
        merge_from.merged_into = int(cuID)
        merge_from.merged_on = datetime.datetime.now()

        # fix email conflict, if any
        if merge_into.email.lower().strip() == merge_from.email.lower().strip():
            merge_from.email = ''

        # if the merge_from account is a teacher, make the merge into account a teacher
        if merge_from.teacher:
            merge_into.teacher = True
        # if the merge_from account is enabled, enable the merge into account
        if merge_from.enabled:
            merge_into.enabled = True
        # if the merge_from account is an employee, make the merge into account an employee
        if merge_from.employee:
            merge_into.employee = True


        merge_from.update_record()
        merge_into.update_record()

        session.flash = T("Merge success")
    else:
        session.flash = T('Merge failed - already merged')


    redirect(URL('edit', args=[cuID]))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('set_password', 'auth_user'))
def account_set_password():
    '''
        Set a new password for an account
    '''
    response.view = 'customers/edit_general.html'
    cuID = request.vars['cuID']
    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T('Set password')

    for field in db.auth_user:
        field.readable = False
        field.writable = False

    db.auth_user.password.readable = True
    db.auth_user.password.writable = True

    form = crud.update(db.auth_user, cuID)

    form_id = "MainForm"
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    submit = form.element('input[type=submit]')

    submenu = account_get_submenu(request.function, cuID)
    description = DIV(B(T("Enter a new password for this account")))

    content = DIV(submenu, BR(),
                  description,
                  form)

    back = edit_get_back()
    menu = customers_get_menu(cuID, request.function)

    return dict(content=content,
                menu=menu,
                back=back,
                save=submit)


def account_get_submenu(page, cuID):
    '''
        Returns submenu for account pages
    '''
    vars = {'cuID':cuID}
    pages = [['account', T('Account'), URL('account', vars=vars)]]

    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('merge', 'auth_user'):
        pages.append(['account_merge', T('Merge'), URL('account_merge',
                                                        vars=vars)])
    if auth.has_membership(group_id='Admins') or \
       auth.has_permission('set_password', 'auth_user'):
        pages.append(['account_set_password', T('Reset password'),
                       URL('account_set_password', vars=vars)])

    horizontal = True

    return os_gui.get_submenu(pages, page, horizontal=horizontal, htype='tabs')


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'invoices'))
def invoices():
    '''
        List invoices for a customer
    '''
    cuID = request.vars['cuID']

    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Invoices")
    response.view = 'customers/edit_general.html'

    session.invoices_edit_back = 'customers_invoices'
    session.invoices_payment_add_back = 'customers_invoices'
    # Always reset filter
    session.invoices_list_status = None

    ih = InvoicesHelper()
    form = ih.add_get_form(cuID)
    result = ih.add_get_modal(form)
    add = result['button']
    modal_class = result['modal_class']

    status_filter = ih.list_get_status_filter()

    content = DIV(status_filter, BR())

    ih = InvoicesHelper()
    list = ih.list_invoices(cuID=cuID)
    content.append(list)

    back = edit_get_back()
    menu = customers_get_menu(cuID, request.function)

    return dict(content=content,
                menu=menu,
                back=back,
                form_add=form,
                add=add,
                modal_class=modal_class,
                left_sidebar_enabled=True)




@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'customers_orders'))
def orders():
    '''
        List orders for a customer
    '''
    cuID = request.vars['cuID']

    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Orders")
    response.view = 'customers/edit_general.html'

    session.invoices_edit_back = 'customers_orders'
    session.invoices_payment_add_back = 'customers_orders'
    session.orders_edit_back = 'customers_orders'

    header = THEAD(TR(TH(T('Order #')),
                      TH(T('Time')),
                      TH(T('Total amount incl. VAT')),
                      TH(T('Status')),
                      TH(T('Invoice')),
                      TH())
                     )

    table = TABLE(header, _class='table table-striped table-hover')
    rows = customer.get_orders_rows()
    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        link_view = os_gui.get_button('edit', URL('orders', 'edit', vars={'cuID': cuID,
                                                                          'coID': row.customers_orders.id}),
                                      _class='pull-right',
                                      btn_size='')

        table.append(TR(TD(repr_row.customers_orders.id),
                        TD(repr_row.customers_orders.DateCreated),
                        TD(repr_row.customers_orders_amounts.TotalPriceVAT),
                        TD(repr_row.customers_orders.Status),
                        TD(orders_get_link_invoice(row)),
                        TD(link_view)
                        ))

    content = table

    back = edit_get_back()
    menu = customers_get_menu(cuID, request.function)

    return dict(content=content,
                menu=menu,
                back=back,
                left_sidebar_enabled=True)


def orders_get_link_invoice(row):
    '''
        Returns invoice for an order in list
    '''
    if row.invoices.id:
        ih = InvoicesHelper()

        query = (db.invoices.id == row.invoices.id)
        rows = db(query).select(db.invoices.ALL)
        repr_row = rows.render(0)

        invoice_link = ih.represent_invoice_for_list(
            row.invoices.id,
            repr_row.InvoiceID,
            repr_row.Status,
            row.invoices.Status,
            row.invoices.payment_methods_id
        )

    else:
        invoice_link = ''

    return invoice_link



@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('read', 'customers_orders'))
def order():
    '''
        Display order content for a customer
    '''
    cuID = request.vars['cuID']
    coID = request.vars['coID']

    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Order #") + coID
    response.view = 'customers/edit_general.html'

    content = DIV(os_gui.get_button('back_bs', URL('orders', vars={'cuID': cuID})),
                  BR(), BR(),
                  DIV(LOAD('orders', 'display_order.load', ajax=False,
                           vars={'coID': coID}))
                  )

    back = edit_get_back()
    menu = customers_get_menu(cuID, 'orders')

    return dict(content=content,
                menu=menu,
                back=back,
                left_sidebar_enabled=True)


@auth.requires(auth.has_membership(group_id='Admins') or \
                auth.has_permission('update', 'auth_user'))
def edit_teacher():
    '''
        Teacher profile for a user
    '''
    cuID = request.vars['cuID']
    response.view = 'customers/edit_general.html'

    customer = Customer(cuID)
    response.title = customer.get_name()
    response.subtitle = T("Teacher profile")

    # disable requires for all fields to allow submitting a limited nr of values
    for field in db.auth_user:
        field.readable=False
        field.writable=False

    fields = [ db.auth_user.teacher_role,
               db.auth_user.teacher_bio,
               db.auth_user.education,
               db.auth_user.teacher_bio_link,
               db.auth_user.teacher_website ]

    for field in fields:
        field.readable=True
        field.writable=True

    crud.messages.submit_button = T('Save')
    crud.messages.record_updated = T('Saved')
    crud.settings.update_onaccept = [cache_clear_school_teachers, cache_clear_classschedule]
    form = crud.update(db.auth_user, cuID)

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    textareas = form.elements('textarea')
    for textarea in textareas:
        textarea['_class'] += ' tmced'


    teacher_info = DIV(
        XML('<form action="#" id="MainForm" enctype="multipart/form-data" method="post">'),
        #form.custom.begin,
        # INPUT(_name='email',
        #       _value=customer.row.email,
        #       _type='hidden',
        #       ),
        DIV(DIV(LABEL(form.custom.label.teacher_role),
                form.custom.widget.teacher_role,
                _class='col-md-12'),
            DIV(LABEL(form.custom.label.teacher_bio),
                form.custom.widget.teacher_bio,
                _class='col-md-6'),
            DIV(LABEL(form.custom.label.education),
                form.custom.widget.education,
                _class='col-md-6'),
            DIV(LABEL(form.custom.label.teacher_bio_link),
                    form.custom.widget.teacher_bio_link,
                    _class='col-md-6'),
            DIV(LABEL(form.custom.label.teacher_website),
                form.custom.widget.teacher_website,
                _class='col-md-6'),
            _class='row'),
        form.custom.end,
        _class='tab-pane active')

    back = edit_get_back()
    menu = customers_get_menu(cuID, request.function)

    return dict(content=teacher_info,
                menu=menu,
                back=back,
                save=submit,
                left_sidebar_enabled=True)